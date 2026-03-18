"""
MKJ SUPA CUP Appeals — Jury Chair Views
Provides read-only data views + PDF / Excel exports for the Chair of the Jury.

Data sections:
  1. Teams
  2. Approved players (per team)
  3. Fixtures
  4. Match reports
  5. Match day squads & starting eleven
  6. Disciplinary sanctions (yellow / red cards)
"""
import io
from datetime import datetime, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Count, Sum, F

from accounts.models import UserRole
from teams.models import Team, Player
from competitions.models import Competition, Fixture
from matches.models import (
    MatchReport, MatchEvent, SquadSubmission, SquadPlayer,
)


# ─── Access guard ──────────────────────────────────────────────────────────────
def jury_chair_required(view_func):
    """Only allow jury_chair role or admin."""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("web_login")
        if request.user.role != UserRole.JURY_CHAIR and not request.user.is_admin:
            messages.error(request, "Only the Chair of the Jury can access this page.")
            return redirect("dashboard")
        return view_func(request, *args, **kwargs)
    return wrapper


# ══════════════════════════════════════════════════════════════════════════════
#  1. TEAMS
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@jury_chair_required
def jury_teams_view(request):
    """List all registered teams with filter support."""
    teams = Team.objects.select_related("competition", "manager", "county").all()

    competition_filter = request.GET.get("competition", "")
    county_filter = request.GET.get("county", "")
    status_filter = request.GET.get("status", "")
    search = request.GET.get("search", "")

    if competition_filter:
        teams = teams.filter(competition_id=competition_filter)
    if county_filter:
        teams = teams.filter(county=county_filter)
    if status_filter:
        teams = teams.filter(status=status_filter)
    if search:
        teams = teams.filter(Q(name__icontains=search) | Q(contact_email__icontains=search))

    competitions = Competition.objects.all().order_by("name")
    return render(request, "jury/teams.html", {
        "teams": teams.order_by("name"),
        "competitions": competitions,
        "filters": {
            "competition": competition_filter,
            "county": county_filter,
            "status": status_filter,
            "search": search,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
#  2. APPROVED PLAYERS
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@jury_chair_required
def jury_players_view(request):
    """List approved/verified players, optionally filtered by team."""
    players = Player.objects.select_related("team", "team__competition").all()

    team_filter = request.GET.get("team", "")
    competition_filter = request.GET.get("competition", "")
    status_filter = request.GET.get("verification", "")
    search = request.GET.get("search", "")

    if team_filter:
        players = players.filter(team_id=team_filter)
    if competition_filter:
        players = players.filter(team__competition_id=competition_filter)
    if status_filter:
        players = players.filter(verification_status=status_filter)
    else:
        # Default: show only verified players
        players = players.filter(verification_status="verified")
    if search:
        players = players.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(national_id_number__icontains=search)
        )

    teams = Team.objects.filter(status="registered").order_by("name")
    competitions = Competition.objects.all().order_by("name")
    return render(request, "jury/players.html", {
        "players": players.order_by("team__name", "last_name"),
        "teams": teams,
        "competitions": competitions,
        "filters": {
            "team": team_filter,
            "competition": competition_filter,
            "verification": status_filter,
            "search": search,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
#  3. FIXTURES
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@jury_chair_required
def jury_fixtures_view(request):
    """All fixtures across competitions."""
    fixtures = Fixture.objects.select_related(
        "competition", "home_team", "away_team", "venue", "pool",
    ).all()

    competition_filter = request.GET.get("competition", "")
    status_filter = request.GET.get("status", "")
    search = request.GET.get("search", "")

    if competition_filter:
        fixtures = fixtures.filter(competition_id=competition_filter)
    if status_filter:
        fixtures = fixtures.filter(status=status_filter)
    if search:
        fixtures = fixtures.filter(
            Q(home_team__name__icontains=search) | Q(away_team__name__icontains=search)
        )

    competitions = Competition.objects.all().order_by("name")
    return render(request, "jury/fixtures.html", {
        "fixtures": fixtures.order_by("-match_date", "-kickoff_time"),
        "competitions": competitions,
        "filters": {
            "competition": competition_filter,
            "status": status_filter,
            "search": search,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
#  4. MATCH REPORTS
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@jury_chair_required
def jury_match_reports_view(request):
    """All submitted/approved match reports."""
    reports = MatchReport.objects.select_related(
        "fixture", "fixture__competition",
        "fixture__home_team", "fixture__away_team",
        "referee", "referee__user",
    ).all()

    competition_filter = request.GET.get("competition", "")
    status_filter = request.GET.get("status", "")
    search = request.GET.get("search", "")

    if competition_filter:
        reports = reports.filter(fixture__competition_id=competition_filter)
    if status_filter:
        reports = reports.filter(status=status_filter)
    if search:
        reports = reports.filter(
            Q(fixture__home_team__name__icontains=search)
            | Q(fixture__away_team__name__icontains=search)
        )

    competitions = Competition.objects.all().order_by("name")
    return render(request, "jury/match_reports.html", {
        "reports": reports.order_by("-submitted_at"),
        "competitions": competitions,
        "filters": {
            "competition": competition_filter,
            "status": status_filter,
            "search": search,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
#  5. MATCH DAY SQUADS & STARTING ELEVEN
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@jury_chair_required
def jury_squads_view(request):
    """List all squad submissions."""
    squads = SquadSubmission.objects.select_related(
        "fixture", "fixture__competition",
        "fixture__home_team", "fixture__away_team", "team",
    ).all()

    competition_filter = request.GET.get("competition", "")
    team_filter = request.GET.get("team", "")
    status_filter = request.GET.get("status", "")

    if competition_filter:
        squads = squads.filter(fixture__competition_id=competition_filter)
    if team_filter:
        squads = squads.filter(team_id=team_filter)
    if status_filter:
        squads = squads.filter(status=status_filter)

    competitions = Competition.objects.all().order_by("name")
    teams = Team.objects.filter(status="registered").order_by("name")
    return render(request, "jury/squads.html", {
        "squads": squads.order_by("-fixture__match_date"),
        "competitions": competitions,
        "teams": teams,
        "filters": {
            "competition": competition_filter,
            "team": team_filter,
            "status": status_filter,
        },
    })


@login_required
@jury_chair_required
def jury_squad_detail_view(request, pk):
    """View a single squad submission with starters and subs."""
    squad = get_object_or_404(
        SquadSubmission.objects.select_related(
            "fixture", "fixture__competition",
            "fixture__home_team", "fixture__away_team", "team",
        ),
        pk=pk,
    )
    players = SquadPlayer.objects.filter(submission=squad).select_related("player").order_by(
        "-is_starter", "shirt_number"
    )
    starters = players.filter(is_starter=True)
    subs = players.filter(is_starter=False)
    return render(request, "jury/squad_detail.html", {
        "squad": squad,
        "starters": starters,
        "subs": subs,
        "all_players": players,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  6. DISCIPLINARY SANCTIONS
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@jury_chair_required
def jury_disciplinary_view(request):
    """
    Disciplinary overview: yellow cards, red cards, second yellows per team.
    """
    card_types = ["yellow_card", "red_card", "second_yellow"]
    events = MatchEvent.objects.filter(
        event_type__in=card_types,
    ).select_related(
        "report", "report__fixture", "report__fixture__competition",
        "team", "player",
    )

    competition_filter = request.GET.get("competition", "")
    team_filter = request.GET.get("team", "")
    card_filter = request.GET.get("card_type", "")
    search = request.GET.get("search", "")

    if competition_filter:
        events = events.filter(report__fixture__competition_id=competition_filter)
    if team_filter:
        events = events.filter(team_id=team_filter)
    if card_filter:
        events = events.filter(event_type=card_filter)
    if search:
        events = events.filter(
            Q(player__first_name__icontains=search)
            | Q(player__last_name__icontains=search)
            | Q(team__name__icontains=search)
        )

    # Summary per team
    team_summary = (
        MatchEvent.objects.filter(event_type__in=card_types)
        .values("team__id", "team__name")
        .annotate(
            yellows=Count("id", filter=Q(event_type="yellow_card")),
            reds=Count("id", filter=Q(event_type="red_card")),
            second_yellows=Count("id", filter=Q(event_type="second_yellow")),
            total=Count("id"),
        )
        .order_by("-total")
    )
    if competition_filter:
        team_summary = team_summary.filter(report__fixture__competition_id=competition_filter)

    competitions = Competition.objects.all().order_by("name")
    teams = Team.objects.filter(status="registered").order_by("name")
    return render(request, "jury/disciplinary.html", {
        "events": events.order_by("-report__fixture__match_date", "-minute"),
        "team_summary": team_summary,
        "competitions": competitions,
        "teams": teams,
        "filters": {
            "competition": competition_filter,
            "team": team_filter,
            "card_type": card_filter,
            "search": search,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORTS
# ══════════════════════════════════════════════════════════════════════════════

def _xl_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        "header_font": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
        "header_fill": PatternFill(start_color="004D1A", end_color="004D1A", fill_type="solid"),
        "title_font": Font(name="Calibri", bold=True, size=14, color="004D1A"),
        "subtitle_font": Font(name="Calibri", size=10, color="666666"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "center": Alignment(horizontal="center", vertical="center"),
        "wrap": Alignment(vertical="center", wrap_text=True),
    }


def _xl_write_header(ws, title, subtitle, headers, col_widths, styles):
    from openpyxl.styles import Alignment
    last_col_letter = chr(64 + len(headers))
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font = styles["title_font"]
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = subtitle
    ws["A2"].font = styles["subtitle_font"]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["thin_border"]
        ws.column_dimensions[cell.column_letter].width = w


def _xl_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ── Teams Excel ───────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_teams_excel(request):
    from openpyxl import Workbook
    styles = _xl_styles()
    now = timezone.localtime(timezone.now())

    teams = Team.objects.select_related("competition", "manager").order_by("name")
    comp_filter = request.GET.get("competition", "")
    if comp_filter:
        teams = teams.filter(competition_id=comp_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"
    headers = ["#", "Team Name", "County", "Competition", "Status", "Manager", "Contact Email", "Contact Phone", "Registered"]
    widths = [6, 30, 18, 30, 14, 30, 30, 18, 18]
    _xl_write_header(ws, "MKJ SUPA CUP — Teams Report", f"Generated: {now.strftime('%B %d, %Y %I:%M %p')}  |  Records: {teams.count()}", headers, widths, styles)

    for i, t in enumerate(teams[:5000], start=5):
        vals = [
            i - 4, t.name, t.county or "-",
            t.competition.name if t.competition else "-",
            t.get_status_display() if hasattr(t, "get_status_display") else t.status,
            t.manager.get_full_name() if t.manager else "-",
            t.contact_email or "-", t.contact_phone or "-",
            t.registered_at.strftime("%Y-%m-%d") if t.registered_at else "-",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = styles["thin_border"]

    return _xl_response(wb, f"MKJ SUPA CUP_Teams_{now.strftime('%Y%m%d_%H%M')}.xlsx")


# ── Players Excel ─────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_players_excel(request):
    from openpyxl import Workbook
    styles = _xl_styles()
    now = timezone.localtime(timezone.now())

    players = Player.objects.select_related("team", "team__competition").filter(verification_status="verified").order_by("team__name", "last_name")
    team_filter = request.GET.get("team", "")
    comp_filter = request.GET.get("competition", "")
    if team_filter:
        players = players.filter(team_id=team_filter)
    if comp_filter:
        players = players.filter(team__competition_id=comp_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Players"
    headers = ["#", "First Name", "Last Name", "Team", "Competition", "Position", "Shirt #", "DOB", "Age", "ID Number", "Status"]
    widths = [6, 20, 20, 28, 28, 12, 10, 14, 6, 18, 14]
    _xl_write_header(ws, "MKJ SUPA CUP — Approved Players Report", f"Generated: {now.strftime('%B %d, %Y %I:%M %p')}  |  Records: {players.count()}", headers, widths, styles)

    for i, p in enumerate(players[:5000], start=5):
        vals = [
            i - 4, p.first_name, p.last_name, p.team.name,
            p.team.competition.name if p.team.competition else "-",
            p.get_position_display() if hasattr(p, "get_position_display") else p.position,
            p.shirt_number,
            p.date_of_birth.strftime("%Y-%m-%d") if p.date_of_birth else "-",
            p.age if hasattr(p, "age") else "-",
            p.national_id_number or "-",
            p.get_status_display() if hasattr(p, "get_status_display") else p.status,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = styles["thin_border"]

    return _xl_response(wb, f"MKJ SUPA CUP_Approved_Players_{now.strftime('%Y%m%d_%H%M')}.xlsx")


# ── Fixtures Excel ────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_fixtures_excel(request):
    from openpyxl import Workbook
    styles = _xl_styles()
    now = timezone.localtime(timezone.now())

    fixtures = Fixture.objects.select_related("competition", "home_team", "away_team", "venue").order_by("-match_date")
    comp_filter = request.GET.get("competition", "")
    if comp_filter:
        fixtures = fixtures.filter(competition_id=comp_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fixtures"
    headers = ["#", "Competition", "Home Team", "Away Team", "Date", "Kickoff", "Venue", "Status", "Score"]
    widths = [6, 28, 25, 25, 14, 10, 28, 14, 12]
    _xl_write_header(ws, "MKJ SUPA CUP — Fixtures Report", f"Generated: {now.strftime('%B %d, %Y %I:%M %p')}  |  Records: {fixtures.count()}", headers, widths, styles)

    for i, f in enumerate(fixtures[:5000], start=5):
        score = ""
        if f.status == "completed":
            score = f"{f.home_score} - {f.away_score}"
        vals = [
            i - 4,
            f.competition.name if f.competition else "-",
            f.home_team.name if f.home_team else "TBD",
            f.away_team.name if f.away_team else "TBD",
            f.match_date.strftime("%Y-%m-%d") if f.match_date else "-",
            f.kickoff_time.strftime("%H:%M") if f.kickoff_time else "-",
            f.venue.name if f.venue else "-",
            f.get_status_display() if hasattr(f, "get_status_display") else f.status,
            score,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = styles["thin_border"]

    return _xl_response(wb, f"MKJ SUPA CUP_Fixtures_{now.strftime('%Y%m%d_%H%M')}.xlsx")


# ── Match Reports Excel ──────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_match_reports_excel(request):
    from openpyxl import Workbook
    styles = _xl_styles()
    now = timezone.localtime(timezone.now())

    reports = MatchReport.objects.select_related(
        "fixture", "fixture__competition", "fixture__home_team",
        "fixture__away_team", "referee", "referee__user",
    ).order_by("-submitted_at")
    comp_filter = request.GET.get("competition", "")
    if comp_filter:
        reports = reports.filter(fixture__competition_id=comp_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Match Reports"
    headers = ["#", "Competition", "Home Team", "Away Team", "Score", "Yellow (H)", "Yellow (A)", "Red (H)", "Red (A)", "Referee", "Status", "Submitted"]
    widths = [6, 28, 22, 22, 10, 10, 10, 10, 10, 28, 14, 18]
    _xl_write_header(ws, "MKJ SUPA CUP — Match Reports", f"Generated: {now.strftime('%B %d, %Y %I:%M %p')}  |  Records: {reports.count()}", headers, widths, styles)

    for i, r in enumerate(reports[:5000], start=5):
        fx = r.fixture
        vals = [
            i - 4,
            fx.competition.name if fx.competition else "-",
            fx.home_team.name if fx.home_team else "TBD",
            fx.away_team.name if fx.away_team else "TBD",
            f"{r.home_score} - {r.away_score}",
            r.home_yellow_cards or 0, r.away_yellow_cards or 0,
            r.home_red_cards or 0, r.away_red_cards or 0,
            r.referee.user.get_full_name() if r.referee and r.referee.user else "-",
            r.get_status_display() if hasattr(r, "get_status_display") else r.status,
            r.submitted_at.strftime("%Y-%m-%d %H:%M") if r.submitted_at else "-",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = styles["thin_border"]

    return _xl_response(wb, f"MKJ SUPA CUP_Match_Reports_{now.strftime('%Y%m%d_%H%M')}.xlsx")


# ── Squads Excel ──────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_squads_excel(request):
    from openpyxl import Workbook
    styles = _xl_styles()
    now = timezone.localtime(timezone.now())

    squad_players = SquadPlayer.objects.select_related(
        "submission", "submission__fixture", "submission__fixture__competition",
        "submission__team", "player",
    ).order_by("submission__fixture__match_date", "submission__team__name", "-is_starter", "shirt_number")

    comp_filter = request.GET.get("competition", "")
    team_filter = request.GET.get("team", "")
    if comp_filter:
        squad_players = squad_players.filter(submission__fixture__competition_id=comp_filter)
    if team_filter:
        squad_players = squad_players.filter(submission__team_id=team_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Squads"
    headers = ["#", "Match Date", "Competition", "Team", "vs", "Player", "Shirt #", "Position", "Starter", "Status"]
    widths = [6, 14, 28, 25, 25, 28, 10, 12, 10, 14]
    _xl_write_header(ws, "MKJ SUPA CUP — Match Day Squads", f"Generated: {now.strftime('%B %d, %Y %I:%M %p')}  |  Records: {squad_players.count()}", headers, widths, styles)

    for i, sp in enumerate(squad_players[:5000], start=5):
        sub = sp.submission
        fx = sub.fixture
        opponent = fx.away_team.name if sub.team == fx.home_team else fx.home_team.name
        vals = [
            i - 4,
            fx.match_date.strftime("%Y-%m-%d") if fx.match_date else "-",
            fx.competition.name if fx.competition else "-",
            sub.team.name,
            opponent if opponent else "TBD",
            f"{sp.player.first_name} {sp.player.last_name}" if sp.player else "-",
            sp.shirt_number or "-",
            sp.player.get_position_display() if sp.player and hasattr(sp.player, "get_position_display") else "-",
            "Yes" if sp.is_starter else "No",
            sub.get_status_display() if hasattr(sub, "get_status_display") else sub.status,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = styles["thin_border"]

    return _xl_response(wb, f"MKJ SUPA CUP_Squads_{now.strftime('%Y%m%d_%H%M')}.xlsx")


# ── Disciplinary Excel ────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_disciplinary_excel(request):
    from openpyxl import Workbook
    styles = _xl_styles()
    now = timezone.localtime(timezone.now())

    card_types = ["yellow_card", "red_card", "second_yellow"]
    events = MatchEvent.objects.filter(event_type__in=card_types).select_related(
        "report", "report__fixture", "report__fixture__competition", "team", "player",
    ).order_by("-report__fixture__match_date", "-minute")

    comp_filter = request.GET.get("competition", "")
    team_filter = request.GET.get("team", "")
    if comp_filter:
        events = events.filter(report__fixture__competition_id=comp_filter)
    if team_filter:
        events = events.filter(team_id=team_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Disciplinary"
    headers = ["#", "Match Date", "Competition", "Team", "Player", "Card Type", "Minute", "Match", "Notes"]
    widths = [6, 14, 28, 25, 28, 16, 10, 40, 30]
    _xl_write_header(ws, "MKJ SUPA CUP — Disciplinary Sanctions Report", f"Generated: {now.strftime('%B %d, %Y %I:%M %p')}  |  Records: {events.count()}", headers, widths, styles)

    for i, ev in enumerate(events[:5000], start=5):
        fx = ev.report.fixture
        vals = [
            i - 4,
            fx.match_date.strftime("%Y-%m-%d") if fx.match_date else "-",
            fx.competition.name if fx.competition else "-",
            ev.team.name if ev.team else "-",
            f"{ev.player.first_name} {ev.player.last_name}" if ev.player else "Unknown",
            ev.get_event_type_display() if hasattr(ev, "get_event_type_display") else ev.event_type,
            ev.minute or "-",
            f"{fx.home_team.name} vs {fx.away_team.name}" if fx.home_team and fx.away_team else "-",
            ev.notes or "-",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = styles["thin_border"]

    return _xl_response(wb, f"MKJ SUPA CUP_Disciplinary_{now.strftime('%Y%m%d_%H%M')}.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORTS  (reportlab)
# ══════════════════════════════════════════════════════════════════════════════

def _pdf_base(title, subtitle, header_row, data_rows, col_widths, filename):
    """Generic PDF builder following the existing MKJ SUPA CUP export style with 11th edition logo."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from django.conf import settings
    import os
    from django.conf import settings
    import os

    now = timezone.localtime(timezone.now())
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#004D1A"), spaceAfter=4)
    subtitle_style = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.gray, spaceAfter=12)
    cell_style = ParagraphStyle("C", parent=styles["Normal"], fontSize=7.5, leading=10)

    # Add MKJ SUPA CUP 11th Edition logo at the top
    elements = []
    logo_path = os.path.join(settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else settings.STATIC_ROOT, 'img', 'mkj_supacup_logo_official.jpg')
    if os.path.exists(logo_path):
        try:
            logo = Image(logo_path, width=40*mm, height=40*mm)
            logo.hAlign = 'LEFT'
            elements.append(logo)
            elements.append(Spacer(1, 2*mm))
        except:
            pass  # Logo load failed, continue without it
    
    # Add MKJ SUPA CUP 11th Edition header
    elements.extend([
        Paragraph("⚽ KENYA YOUTH INTERCOUNTY SPORTS ASSOCIATION — 11TH EDITION", 
                 ParagraphStyle("Edition", parent=styles["Heading2"], fontSize=10, textColor=colors.HexColor("#004D1A"), spaceAfter=2, alignment=1)),
        Paragraph(title, title_style),
        Paragraph(subtitle, ParagraphStyle("S2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#333"), spaceAfter=4)),
        Paragraph(f"Generated: {now.strftime('%B %d, %Y at %I:%M %p')}  &bull;  Records: {len(data_rows)}", subtitle_style),
        Spacer(1, 6 * mm),
    ])

    table_data = [header_row]
    for row in data_rows[:2000]:
        table_data.append([
            Paragraph(str(c), cell_style) if len(str(c)) > 25 else str(c) for c in row
        ])

    mkj_supacup_green = colors.HexColor("#004D1A")
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), mkj_supacup_green),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F7F0")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph(
        f"<i>Report generated by MKJ SUPA CUP CMS on {now.strftime('%d/%m/%Y %H:%M')}. Confidential — for authorised personnel only.</i>",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.gray),
    ))

    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ── Teams PDF ─────────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_teams_pdf(request):
    now = timezone.localtime(timezone.now())
    teams = Team.objects.select_related("competition", "manager").order_by("name")
    comp_filter = request.GET.get("competition", "")
    if comp_filter:
        teams = teams.filter(competition_id=comp_filter)

    header = ["#", "Team", "County", "Competition", "Status", "Manager", "Email", "Phone"]
    rows = []
    for i, t in enumerate(teams, 1):
        rows.append([
            str(i), t.name, t.county or "-",
            t.competition.name if t.competition else "-",
            t.get_status_display() if hasattr(t, "get_status_display") else t.status,
            t.manager.get_full_name() if t.manager else "-",
            t.contact_email or "-", t.contact_phone or "-",
        ])
    return _pdf_base(
        "MKJ SUPA CUP Competition Management System", "Teams Report",
        header, rows, [22, 80, 55, 80, 45, 70, 90, 60],
        f"MKJ SUPA CUP_Teams_{now.strftime('%Y%m%d_%H%M')}.pdf",
    )


# ── Players PDF ───────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_players_pdf(request):
    now = timezone.localtime(timezone.now())
    players = Player.objects.select_related("team", "team__competition").filter(
        verification_status="verified"
    ).order_by("team__name", "last_name")
    team_filter = request.GET.get("team", "")
    comp_filter = request.GET.get("competition", "")
    if team_filter:
        players = players.filter(team_id=team_filter)
    if comp_filter:
        players = players.filter(team__competition_id=comp_filter)

    header = ["#", "Name", "Team", "Competition", "Position", "Shirt", "DOB", "Age", "ID"]
    rows = []
    for i, p in enumerate(players, 1):
        rows.append([
            str(i), f"{p.first_name} {p.last_name}", p.team.name,
            p.team.competition.name if p.team.competition else "-",
            p.get_position_display() if hasattr(p, "get_position_display") else p.position,
            str(p.shirt_number), p.date_of_birth.strftime("%Y-%m-%d") if p.date_of_birth else "-",
            str(p.age) if hasattr(p, "age") else "-", p.national_id_number or "-",
        ])
    return _pdf_base(
        "MKJ SUPA CUP Competition Management System", "Approved Players Report",
        header, rows, [22, 75, 70, 70, 45, 30, 50, 25, 60],
        f"MKJ SUPA CUP_Players_{now.strftime('%Y%m%d_%H%M')}.pdf",
    )


# ── Fixtures PDF ──────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_fixtures_pdf(request):
    now = timezone.localtime(timezone.now())
    fixtures = Fixture.objects.select_related("competition", "home_team", "away_team", "venue").order_by("-match_date")
    comp_filter = request.GET.get("competition", "")
    if comp_filter:
        fixtures = fixtures.filter(competition_id=comp_filter)

    header = ["#", "Competition", "Home", "Away", "Date", "Kickoff", "Venue", "Status", "Score"]
    rows = []
    for i, f in enumerate(fixtures, 1):
        score = f"{f.home_score} - {f.away_score}" if f.status == "completed" else "-"
        rows.append([
            str(i), f.competition.name if f.competition else "-",
            f.home_team.name if f.home_team else "TBD",
            f.away_team.name if f.away_team else "TBD",
            f.match_date.strftime("%Y-%m-%d") if f.match_date else "-",
            f.kickoff_time.strftime("%H:%M") if f.kickoff_time else "-",
            f.venue.name if f.venue else "-",
            f.get_status_display() if hasattr(f, "get_status_display") else f.status,
            score,
        ])
    return _pdf_base(
        "MKJ SUPA CUP Competition Management System", "Fixtures Report",
        header, rows, [22, 75, 65, 65, 50, 40, 75, 50, 40],
        f"MKJ SUPA CUP_Fixtures_{now.strftime('%Y%m%d_%H%M')}.pdf",
    )


# ── Match Reports PDF ────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_match_reports_pdf(request):
    now = timezone.localtime(timezone.now())
    reports = MatchReport.objects.select_related(
        "fixture", "fixture__competition", "fixture__home_team",
        "fixture__away_team", "referee", "referee__user",
    ).order_by("-submitted_at")
    comp_filter = request.GET.get("competition", "")
    if comp_filter:
        reports = reports.filter(fixture__competition_id=comp_filter)

    header = ["#", "Competition", "Home", "Away", "Score", "YC (H)", "YC (A)", "RC (H)", "RC (A)", "Referee", "Status"]
    rows = []
    for i, r in enumerate(reports, 1):
        fx = r.fixture
        rows.append([
            str(i), fx.competition.name if fx.competition else "-",
            fx.home_team.name if fx.home_team else "TBD",
            fx.away_team.name if fx.away_team else "TBD",
            f"{r.home_score}-{r.away_score}",
            str(r.home_yellow_cards or 0), str(r.away_yellow_cards or 0),
            str(r.home_red_cards or 0), str(r.away_red_cards or 0),
            r.referee.user.get_full_name() if r.referee and r.referee.user else "-",
            r.get_status_display() if hasattr(r, "get_status_display") else r.status,
        ])
    return _pdf_base(
        "MKJ SUPA CUP Competition Management System", "Match Reports",
        header, rows, [22, 70, 55, 55, 35, 30, 30, 30, 30, 65, 45],
        f"MKJ SUPA CUP_Match_Reports_{now.strftime('%Y%m%d_%H%M')}.pdf",
    )


# ── Squads PDF ────────────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_squads_pdf(request):
    now = timezone.localtime(timezone.now())
    squad_players = SquadPlayer.objects.select_related(
        "submission", "submission__fixture", "submission__fixture__competition",
        "submission__team", "player",
    ).order_by("submission__fixture__match_date", "submission__team__name", "-is_starter", "shirt_number")
    comp_filter = request.GET.get("competition", "")
    team_filter = request.GET.get("team", "")
    if comp_filter:
        squad_players = squad_players.filter(submission__fixture__competition_id=comp_filter)
    if team_filter:
        squad_players = squad_players.filter(submission__team_id=team_filter)

    header = ["#", "Date", "Competition", "Team", "Opponent", "Player", "Shirt", "Starter"]
    rows = []
    for i, sp in enumerate(squad_players, 1):
        sub = sp.submission
        fx = sub.fixture
        opp = fx.away_team.name if sub.team == fx.home_team else (fx.home_team.name if fx.home_team else "TBD")
        rows.append([
            str(i),
            fx.match_date.strftime("%Y-%m-%d") if fx.match_date else "-",
            fx.competition.name if fx.competition else "-",
            sub.team.name,
            opp,
            f"{sp.player.first_name} {sp.player.last_name}" if sp.player else "-",
            str(sp.shirt_number or "-"),
            "Yes" if sp.is_starter else "No",
        ])
    return _pdf_base(
        "MKJ SUPA CUP Competition Management System", "Match Day Squads & Starting Eleven",
        header, rows, [22, 50, 70, 70, 70, 80, 30, 35],
        f"MKJ SUPA CUP_Squads_{now.strftime('%Y%m%d_%H%M')}.pdf",
    )


# ── Disciplinary PDF ─────────────────────────────────────────────────────────

@login_required
@jury_chair_required
def jury_export_disciplinary_pdf(request):
    now = timezone.localtime(timezone.now())
    card_types = ["yellow_card", "red_card", "second_yellow"]
    events = MatchEvent.objects.filter(event_type__in=card_types).select_related(
        "report", "report__fixture", "report__fixture__competition", "team", "player",
    ).order_by("-report__fixture__match_date", "-minute")

    comp_filter = request.GET.get("competition", "")
    team_filter = request.GET.get("team", "")
    if comp_filter:
        events = events.filter(report__fixture__competition_id=comp_filter)
    if team_filter:
        events = events.filter(team_id=team_filter)

    header = ["#", "Date", "Competition", "Team", "Player", "Card", "Min", "Match"]
    rows = []
    for i, ev in enumerate(events, 1):
        fx = ev.report.fixture
        rows.append([
            str(i),
            fx.match_date.strftime("%Y-%m-%d") if fx.match_date else "-",
            fx.competition.name if fx.competition else "-",
            ev.team.name if ev.team else "-",
            f"{ev.player.first_name} {ev.player.last_name}" if ev.player else "Unknown",
            ev.get_event_type_display() if hasattr(ev, "get_event_type_display") else ev.event_type,
            str(ev.minute or "-"),
            f"{fx.home_team.name} vs {fx.away_team.name}" if fx.home_team and fx.away_team else "-",
        ])
    return _pdf_base(
        "MKJ SUPA CUP Competition Management System", "Disciplinary Sanctions Report",
        header, rows, [22, 50, 70, 65, 70, 55, 25, 100],
        f"MKJ SUPA CUP_Disciplinary_{now.strftime('%Y%m%d_%H%M')}.pdf",
    )
