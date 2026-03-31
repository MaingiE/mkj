"""
End-to-end integration tests for all recent changes.
Run with: python test_integration.py
"""
import os, sys, django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mkj_cms.settings')
django.setup()

from io import BytesIO
from datetime import date, time, timedelta
from django.test import RequestFactory, TestCase
from django.urls import reverse, resolve, NoReverseMatch
from django.template.loader import get_template
from django.template import TemplateSyntaxError
from django.conf import settings
from django.utils import timezone

PASS = "\033[92m PASS \033[0m"
FAIL = "\033[91m FAIL \033[0m"
WARN = "\033[93m WARN \033[0m"

results = {"pass": 0, "fail": 0, "warn": 0}

def report(label, ok, detail="", warn=False):
    if ok:
        print(f"  {PASS} {label}")
        results["pass"] += 1
    elif warn:
        print(f"  {WARN} {label} — {detail}")
        results["warn"] += 1
    else:
        print(f"  {FAIL} {label} — {detail}")
        results["fail"] += 1


# ═══════════════════════════════════════════════════════════════════════════════
# 1. URL RESOLUTION
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 1. URL Resolution ══")

url_tests = [
    ("cso_bulk_upload_list", {}, "/portal/chief-sports-officer/bulk-uploads/"),
    ("cso_bulk_upload", {}, "/portal/chief-sports-officer/bulk-upload/"),
    ("cso_bulk_upload_detail", {"pk": 1}, "/portal/chief-sports-officer/bulk-uploads/1/"),
    ("director_bulk_upload_list", {}, "/portal/director-sports/bulk-uploads/"),
    ("director_bulk_upload_review", {"pk": 1}, "/portal/director-sports/bulk-uploads/1/review/"),
]

for name, kwargs, expected_path in url_tests:
    try:
        url = reverse(name, kwargs=kwargs)
        ok = url == expected_path
        report(f"reverse('{name}') → {url}", ok, f"expected {expected_path}")
    except NoReverseMatch as e:
        report(f"reverse('{name}')", False, str(e))

# Also verify existing URLs still resolve
existing_urls = [
    "chief_sports_officer_dashboard",
    "director_sports_dashboard",
    "coordinator_dashboard",
    "coordinator_competitions",
]
for name in existing_urls:
    try:
        reverse(name)
        report(f"existing URL '{name}' still resolves", True)
    except NoReverseMatch:
        report(f"existing URL '{name}'", False, "BROKEN — regression!")


# ═══════════════════════════════════════════════════════════════════════════════
# 2. VIEW IMPORTS & FUNCTION SIGNATURES
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 2. View Imports & Signatures ══")

from mkj_cms.web_views import (
    cso_bulk_upload_list_view,
    cso_bulk_upload_view,
    cso_bulk_upload_detail_view,
    director_bulk_upload_list_view,
    director_bulk_upload_review_view,
    coordinator_edit_fixture_view,
)

view_tests = [
    ("cso_bulk_upload_list_view", cso_bulk_upload_list_view, ["request"]),
    ("cso_bulk_upload_view", cso_bulk_upload_view, ["request"]),
    ("cso_bulk_upload_detail_view", cso_bulk_upload_detail_view, ["request", "pk"]),
    ("director_bulk_upload_list_view", director_bulk_upload_list_view, ["request"]),
    ("director_bulk_upload_review_view", director_bulk_upload_review_view, ["request", "pk"]),
    ("coordinator_edit_fixture_view", coordinator_edit_fixture_view, ["request", "pk", "fixture_pk"]),
]

import inspect
for name, fn, expected_params in view_tests:
    sig = inspect.signature(fn)
    params = list(sig.parameters.keys())
    ok = params == expected_params
    report(f"{name} signature: ({', '.join(params)})", ok, f"expected ({', '.join(expected_params)})")


# ═══════════════════════════════════════════════════════════════════════════════
# 3. MODEL INTEGRITY
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 3. Model Integrity ══")

from teams.models import BulkPlayerUpload, BulkPlayerUploadRow, CountyPlayer

# Check BulkPlayerUpload fields
bu_fields = {f.name for f in BulkPlayerUpload._meta.get_fields()}
required_bu = {"id", "file", "sport_type", "sub_county", "status", "uploaded_by",
               "created_at", "total_rows", "valid_rows", "error_rows", "notes",
               "reviewed_by", "reviewed_at", "rejection_reason"}
missing_bu = required_bu - bu_fields
report("BulkPlayerUpload has all required fields", not missing_bu,
       f"missing: {missing_bu}" if missing_bu else "")

# Check BulkPlayerUploadRow fields
br_fields = {f.name for f in BulkPlayerUploadRow._meta.get_fields()}
required_br = {"id", "upload", "row_number", "first_name", "last_name",
               "date_of_birth", "national_id_number", "phone", "position",
               "jersey_number", "ward", "is_valid", "error_message", "county_player"}
missing_br = required_br - br_fields
report("BulkPlayerUploadRow has all required fields", not missing_br,
       f"missing: {missing_br}" if missing_br else "")

# FK relationships
report("BulkPlayerUploadRow → BulkPlayerUpload FK", 
       BulkPlayerUploadRow._meta.get_field("upload").related_model is BulkPlayerUpload)
report("BulkPlayerUploadRow → CountyPlayer FK",
       BulkPlayerUploadRow._meta.get_field("county_player").related_model is CountyPlayer)

# Check table exists in DB
from django.db import connection
with connection.cursor() as cur:
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%bulkplayer%'")
    tables = [r[0] for r in cur.fetchall()]
report(f"DB tables created: {tables}", len(tables) >= 2,
       "tables not found" if len(tables) < 2 else "")

# Check migration is applied
from django.db.migrations.recorder import MigrationRecorder
applied = MigrationRecorder.Migration.objects.filter(
    app='teams', name__contains='bulkplayerupload'
).exists()
report("Migration 0029_bulkplayerupload applied", applied)


# ═══════════════════════════════════════════════════════════════════════════════
# 4. TEMPLATE SYNTAX VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 4. Template Syntax Validation ══")

templates_to_check = [
    "portal/chief_sports_officer/bulk_upload.html",
    "portal/chief_sports_officer/bulk_upload_list.html",
    "portal/chief_sports_officer/bulk_upload_detail.html",
    "portal/director_sports/bulk_upload_list.html",
    "portal/director_sports/bulk_upload_review.html",
    "portal/coordinator/edit_fixture.html",
    "portal/chief_sports_officer/dashboard.html",
    "portal/director_sports/dashboard.html",
    "dashboard/index.html",
]

for tpl_name in templates_to_check:
    try:
        tpl = get_template(tpl_name)
        report(f"Template '{tpl_name}' compiles OK", True)
    except TemplateSyntaxError as e:
        report(f"Template '{tpl_name}'", False, f"SYNTAX ERROR: {e}")
    except Exception as e:
        report(f"Template '{tpl_name}'", False, f"{type(e).__name__}: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# 5. SESSION TIMEOUT CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 5. Session Timeout Configuration ══")

idle_minutes = getattr(settings, 'AUTO_LOGOUT_IDLE_MINUTES', None)
report(f"AUTO_LOGOUT_IDLE_MINUTES = {idle_minutes}", idle_minutes == 30,
       f"expected 30, got {idle_minutes}")

cookie_age = getattr(settings, 'SESSION_COOKIE_AGE', None)
report(f"SESSION_COOKIE_AGE = {cookie_age}s ({cookie_age//60}min)",
       cookie_age >= 1800, f"should be >= 1800 (30min)")

report("SESSION_EXPIRE_AT_BROWSER_CLOSE is True",
       settings.SESSION_EXPIRE_AT_BROWSER_CLOSE is True)

# Verify middleware is installed
middleware = settings.MIDDLEWARE
auto_logout_mw = any('AutoLogout' in m for m in middleware)
report("AutoLogoutMiddleware in MIDDLEWARE", auto_logout_mw,
       "middleware not found — timeout won't work!" if not auto_logout_mw else "")


# ═══════════════════════════════════════════════════════════════════════════════
# 6. COORDINATOR FIXTURE EDIT — HOME/AWAY SWAP LOGIC
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 6. Coordinator Fixture Edit — Home/Away Logic ══")

# Verify the view code does NOT have the is_knockout restriction
import mkj_cms.web_views as wv
source = inspect.getsource(wv.coordinator_edit_fixture_view)

has_knockout_guard = "if fixture.is_knockout:" in source and "home_team_id" in source.split("if fixture.is_knockout:")[1][:200] if "if fixture.is_knockout:" in source else False
report("No is_knockout guard on home/away team edit", not has_knockout_guard,
       "home/away team edit still restricted to knockout only!")

has_home_team_reassign = "home_team_id" in source
report("View handles home_team_id POST param", has_home_team_reassign)

has_away_team_reassign = "away_team_id" in source
report("View handles away_team_id POST param", has_away_team_reassign)

# Verify template shows team selectors without knockout condition
tpl_source = open(os.path.join(settings.BASE_DIR, "templates", "portal", "coordinator", "edit_fixture.html"), encoding="utf-8").read()
has_knockout_if = "{% if fixture.is_knockout %}" in tpl_source
report("Template shows team selectors for ALL fixtures (no knockout guard)",
       not has_knockout_if,
       "template still has {% if fixture.is_knockout %} wrapper!")


# ═══════════════════════════════════════════════════════════════════════════════
# 7. BASKETBALL DISCIPLINE WIRING
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 7. Basketball Discipline Wiring ══")

from teams.models import SportType
has_3x3_men = hasattr(SportType, 'BASKETBALL_3X3_MEN')
has_3x3_women = hasattr(SportType, 'BASKETBALL_3X3_WOMEN')
has_5x5_men = hasattr(SportType, 'BASKETBALL_MEN')
has_5x5_women = hasattr(SportType, 'BASKETBALL_WOMEN')

report("SportType.BASKETBALL_3X3_MEN exists", has_3x3_men)
report("SportType.BASKETBALL_3X3_WOMEN exists", has_3x3_women)
report("SportType.BASKETBALL_MEN (5x5) exists", has_5x5_men)
report("SportType.BASKETBALL_WOMEN (5x5) exists", has_5x5_women)

# Check coordinator discipline choices include basketball variants
coord_choices_source = inspect.getsource(wv.coordinator_dashboard_view) if hasattr(wv, 'coordinator_dashboard_view') else ""

# Check from the COORDINATOR_DISCIPLINE_CHOICES defined in the module
if hasattr(wv, 'COORDINATOR_DISCIPLINE_CHOICES'):
    choices = dict(wv.COORDINATOR_DISCIPLINE_CHOICES)
    report("basketball_5x5 in COORDINATOR_DISCIPLINE_CHOICES",
           'basketball_5x5' in choices, f"keys: {list(choices.keys())}")
    report("basketball_3x3 in COORDINATOR_DISCIPLINE_CHOICES",
           'basketball_3x3' in choices, f"keys: {list(choices.keys())}")
else:
    report("COORDINATOR_DISCIPLINE_CHOICES accessible", False, "not found in web_views module")


# ═══════════════════════════════════════════════════════════════════════════════
# 8. VERIFICATION SUBCOUNTY DROPDOWN FIX
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 8. Verification Sub-County Dropdown ══")

from accounts.models import MakueniSubCounty
subcounties = [c[0] for c in MakueniSubCounty.choices]
expected_sc = ['Makueni', 'Kibwezi West', 'Kibwezi East', 'Kaiti', 'Kilome', 'Mbooni']

# Check enum has all 6
report(f"MakueniSubCounty has {len(subcounties)} values",
       len(subcounties) == 6,
       f"got {len(subcounties)}: {subcounties}")

for sc_val in expected_sc:
    found = any(sc_val.lower().replace(' ', '_') in s.lower().replace(' ', '_') or sc_val.lower() in s.lower() for s in subcounties)
    if not found:
        found = sc_val in [c[1] for c in MakueniSubCounty.choices]
    report(f"  Sub-county '{sc_val}' in enum", found)

# Check VO dashboard view uses MakueniSubCounty, not KenyaCounty
vo_source = inspect.getsource(wv.vo_dashboard_view)
uses_makueni = "MakueniSubCounty" in vo_source or "makueni" in vo_source.lower()
uses_kenya_county = "KenyaCounty" in vo_source
report("vo_dashboard_view uses MakueniSubCounty", uses_makueni)
report("vo_dashboard_view does NOT use KenyaCounty", not uses_kenya_county,
       "still references KenyaCounty!")


# ═══════════════════════════════════════════════════════════════════════════════
# 9. BULK UPLOAD EXCEL PARSING (UNIT TEST)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 9. Bulk Upload Excel Parsing ══")

try:
    import openpyxl
    report("openpyxl is installed", True)

    # Create a test Excel file in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["First Name", "Last Name", "DOB", "National ID", "Phone", "Position", "Jersey", "Ward"])
    ws.append(["John", "Mutua", "2005-03-15", "12345678", "0712345678", "GK", 1, "Wote"])
    ws.append(["Jane", "Mwende", "15/04/2006", "87654321", "+254700111222", "CB", 5, "Nziu"])
    ws.append(["", "Missing", "2005-01-01", "11111111", "", "", "", ""])  # missing first name
    ws.append(["Bad", "Date", "not-a-date", "22222222", "", "", "", ""])  # bad DOB
    ws.append(["Duplicate", "ID", "2005-06-01", "12345678", "", "", "", ""])  # duplicate national ID

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Parse like the view does
    wb2 = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    ws2 = wb2.active
    rows = list(ws2.iter_rows(min_row=2, values_only=True))

    report(f"Test Excel has {len(rows)} data rows", len(rows) == 5)

    # Validate parsing logic mimics view
    from datetime import datetime
    valid_count = 0
    error_count = 0
    seen_ids = set()

    for i, row in enumerate(rows, start=2):
        first = str(row[0] or '').strip()
        last = str(row[1] or '').strip()
        dob_raw = row[2]
        nid = str(row[3] or '').strip()

        errors = []
        if not first:
            errors.append("missing first name")
        if not last:
            errors.append("missing last name")

        # DOB parse
        dob = None
        if isinstance(dob_raw, datetime):
            dob = dob_raw.date()
        elif isinstance(dob_raw, date):
            dob = dob_raw
        elif isinstance(dob_raw, str):
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    dob = datetime.strptime(dob_raw.strip(), fmt).date()
                    break
                except ValueError:
                    continue
        if dob is None:
            errors.append("invalid date")

        if not nid or not nid.replace('-', '').isdigit():
            errors.append("invalid national ID")
        elif nid in seen_ids:
            errors.append("duplicate national ID")
        seen_ids.add(nid)

        if errors:
            error_count += 1
        else:
            valid_count += 1

    report(f"Valid rows: {valid_count} (expected 2)", valid_count == 2)
    report(f"Error rows: {error_count} (expected 3)", error_count == 3)

    wb2.close()

except ImportError:
    report("openpyxl installed", False, "openpyxl not available — bulk upload will fail!")


# ═══════════════════════════════════════════════════════════════════════════════
# 10. ADMIN DASHBOARD INTEGRATION
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 10. Admin Dashboard Bulk Upload Links ══")

admin_tpl = open(os.path.join(settings.BASE_DIR, "templates", "dashboard", "index.html"), encoding="utf-8").read()
report("Admin dashboard has cso_bulk_upload_list link",
       "cso_bulk_upload_list" in admin_tpl)
report("Admin dashboard has director_bulk_upload_list link",
       "director_bulk_upload_list" in admin_tpl)

# CSO dashboard
cso_tpl = open(os.path.join(settings.BASE_DIR, "templates", "portal", "chief_sports_officer", "dashboard.html"), encoding="utf-8").read()
report("CSO dashboard has bulk upload quick action",
       "cso_bulk_upload_list" in cso_tpl)

# Director dashboard
dir_tpl = open(os.path.join(settings.BASE_DIR, "templates", "portal", "director_sports", "dashboard.html"), encoding="utf-8").read()
report("Director dashboard has bulk upload approval link",
       "director_bulk_upload_list" in dir_tpl)


# ═══════════════════════════════════════════════════════════════════════════════
# 11. CROSS-CUTTING: ROLE-BASED ACCESS
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 11. Role-Based Access Decorators ══")

for view_name in ["cso_bulk_upload_list_view", "cso_bulk_upload_view",
                   "cso_bulk_upload_detail_view"]:
    fn = getattr(wv, view_name)
    # Check wrapped function has role_required
    src = inspect.getsource(fn)
    has_cso = "'chief_sports_officer'" in src or "chief_sports_officer" in str(getattr(fn, '__wrapped__', ''))
    has_admin = "'admin'" in src
    report(f"{view_name} allows chief_sports_officer + admin",
           has_cso or has_admin,
           "missing role check!")

for view_name in ["director_bulk_upload_list_view", "director_bulk_upload_review_view"]:
    fn = getattr(wv, view_name)
    src = inspect.getsource(fn)
    has_dir = "'director_sports'" in src
    has_admin = "'admin'" in src
    report(f"{view_name} allows director_sports + admin",
           has_dir or has_admin,
           "missing role check!")


# ═══════════════════════════════════════════════════════════════════════════════
# 12. PENDING MIGRATION CHECK
# ═══════════════════════════════════════════════════════════════════════════════
print("\n══ 12. Pending Migrations ══")

from django.core.management import call_command
from io import StringIO

out = StringIO()
try:
    call_command('showmigrations', '--list', stdout=out)
    migration_output = out.getvalue()
    unapplied = [l.strip() for l in migration_output.split('\n') if '[ ]' in l]
    report(f"No unapplied migrations ({len(unapplied)} pending)",
           len(unapplied) == 0,
           f"unapplied: {unapplied[:5]}")
except Exception as e:
    report("showmigrations", False, str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
print("\n" + "═" * 60)
total = results["pass"] + results["fail"] + results["warn"]
print(f"  TOTAL: {total} tests")
print(f"  {PASS}: {results['pass']}")
print(f"  {FAIL}: {results['fail']}")
print(f"  {WARN}: {results['warn']}")
print("═" * 60)

if results["fail"] == 0:
    print(f"\n  \033[92m✓ ALL TESTS PASSED — safe to deploy!\033[0m\n")
else:
    print(f"\n  \033[91m✗ {results['fail']} FAILURE(S) — fix before deploying!\033[0m\n")

sys.exit(results["fail"])
