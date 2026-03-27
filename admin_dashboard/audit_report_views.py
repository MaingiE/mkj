# admin_dashboard/audit_report_views.py
"""
Comprehensive Audit Report for Super Admin.
Shows who did what, when, across all system actions - with PDF & Excel export.
"""
import io
from collections import OrderedDict
from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.contrib.auth import get_user_model

from .models import ActivityLog

User = get_user_model()


def superadmin_required(user):
    return user.is_superuser or user.role in ('admin', 'secretary_general')


# ── Action category mapping ─────────────────────────────────────────────────
ACTION_CATEGORIES = OrderedDict([
    ("Team Management", [
        "TEAM_CREATE", "TEAM_UPDATE", "TEAM_DELETE",
        "TEAM_APPROVE", "TEAM_REJECT", "TEAM_SUSPEND",
    ]),
    ("Player Management", [
        "PLAYER_CREATE", "PLAYER_UPDATE", "PLAYER_DELETE", "PLAYER_TRANSFER",
    ]),
    ("Payment & Finance", [
        "PAYMENT_RECEIVED", "PAYMENT_VERIFIED", "PAYMENT_ACTION",
    ]),
    ("Match & Fixtures", [
        "MATCH_CREATE", "MATCH_UPDATE", "MATCH_DELETE", "MATCH_RESCHEDULE",
        "MATCH_REPORT", "MATCH_REPORT_APPROVE",
        "FIXTURE_GENERATE", "FIXTURE_REGENERATE",
    ]),
    ("Squad Management", [
        "MATCHDAY_SQUAD_SUBMIT", "SQUAD_APPROVE", "SQUAD_REJECT",
    ]),
    ("Referee Management", [
        "REFEREE_REGISTER", "REFEREE_APPROVE", "REFEREE_ACTION",
    ]),
    ("Competition / Zones", [
        "ZONE_ASSIGN", "ZONE_UPDATE",
    ]),
    ("Suspensions", [
        "SUSPENSION_CREATE", "SUSPENSION_LIFT",
    ]),
    ("User Management", [
        "USER_CREATE", "USER_UPDATE", "USER_DELETE",
        "USER_ROLE_CHANGE", "PASSWORD_CHANGE",
    ]),
    ("Authentication", [
        "LOGIN", "LOGOUT",
    ]),
    ("System & Config", [
        "CONFIG_CHANGE", "REGISTRATION_TOGGLE", "ADMIN_ACTION", "OTHER",
    ]),
    ("Other Actions", []),
])

# Flat reverse lookup
ACTION_TO_CATEGORY = {}
for cat, actions in ACTION_CATEGORIES.items():
    for act in actions:
        ACTION_TO_CATEGORY[act] = cat


def _filtered_audit(request):
    """Filter logs based on request GET parameters."""
    qs = ActivityLog.objects.select_related("user").all()

    category = request.GET.get("category", "")
    user_id = request.GET.get("user", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    search = request.GET.get("search", "")

    # Loosen filter: show all logs if no category is selected
    if category and category in ACTION_CATEGORIES:
        if ACTION_CATEGORIES[category]:
            qs = qs.filter(action__in=ACTION_CATEGORIES[category])
        else:
            # 'Other Actions' category: show logs with actions not in any category
            known_actions = [act for acts in ACTION_CATEGORIES.values() for act in acts]
            qs = qs.exclude(action__in=known_actions)
    # If no filter is set, show all logs (do not filter by action)
    if user_id:
        qs = qs.filter(user_id=user_id)
    if date_from:
        try:
            qs = qs.filter(timestamp__gte=datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(timestamp__lt=datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    if search:
        qs = qs.filter(
            Q(description__icontains=search)
            | Q(object_repr__icontains=search)
            | Q(user__email__icontains=search)
            | Q(user__first_name__icontains=search)
            | Q(user__last_name__icontains=search)
        )
    return qs


# ══════════════════════════════════════════════════════════════════════════════
#  AUDIT REPORT PAGE
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@user_passes_test(superadmin_required)
def audit_report(request):
    """Comprehensive audit report for super admin - who did what and when."""

    logs = _filtered_audit(request)
    now = timezone.localtime(timezone.now())

    # ── Summary statistics ────────────────────────────────────────────────
    total = logs.count()
    today = logs.filter(timestamp__date=now.date()).count()
    this_week = logs.filter(timestamp__gte=now - timedelta(days=7)).count()
    this_month = logs.filter(timestamp__gte=now.replace(day=1)).count()

    # ── Per-category breakdown ────────────────────────────────────────────
    known_actions = [act for acts in ACTION_CATEGORIES.values() for act in acts]
    category_stats = []
    for cat_name, cat_actions in ACTION_CATEGORIES.items():
        if cat_actions:
            count = logs.filter(action__in=cat_actions).count()
        else:
            # 'Other Actions': count logs with actions not in any category
            count = logs.exclude(action__in=known_actions).count()
        if count > 0:
            category_stats.append({"name": cat_name, "count": count})

    # ── Top users by activity ─────────────────────────────────────────────
    user_stats = (
        logs.exclude(user__isnull=True)
        .values("user__id", "user__email", "user__first_name", "user__last_name", "user__role")
        .annotate(count=Count("id"))
        .order_by("-count")[:15]
    )

    # ── Recent critical actions (approvals, payments, suspensions) ────────
    critical_actions = [
        "TEAM_APPROVE", "TEAM_REJECT", "TEAM_SUSPEND",
        "PAYMENT_VERIFIED", "PAYMENT_ACTION",
        "SUSPENSION_CREATE", "SUSPENSION_LIFT",
        "REFEREE_APPROVE",
        "USER_CREATE", "USER_DELETE", "USER_ROLE_CHANGE",
        "FIXTURE_GENERATE",
        "SQUAD_APPROVE", "SQUAD_REJECT",
    ]
    critical_logs = logs.filter(action__in=critical_actions).order_by("-timestamp")[:20]

    # ── All users for filter dropdown ─────────────────────────────────────
    all_users = User.objects.filter(activity_logs__isnull=False).distinct().order_by("email")

    # ── Paginated full log table ──────────────────────────────────────────
    paginator = Paginator(logs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    context = {
        "total": total,
        "today": today,
        "this_week": this_week,
        "this_month": this_month,
        "category_stats": category_stats,
        "user_stats": user_stats,
        "critical_logs": critical_logs,
        "page_obj": page_obj,
        "all_users": all_users,
        "categories": list(ACTION_CATEGORIES.keys()),
        # Current filters (for keeping state)
        "f_category": request.GET.get("category", ""),
        "f_user": request.GET.get("user", ""),
        "f_date_from": request.GET.get("date_from", ""),
        "f_date_to": request.GET.get("date_to", ""),
        "f_search": request.GET.get("search", ""),
        "now": now,
    }
    return render(request, "admin_dashboard/audit_report.html", context)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@user_passes_test(superadmin_required)
def export_audit_excel(request):
    """Export full audit report to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    logs = _filtered_audit(request)
    now = timezone.localtime(timezone.now())

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    # ── Styles ────────────────────────────────────────────────────────────
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="004D1A", end_color="004D1A", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="004D1A")
    sub_font = Font(name="Calibri", size=10, color="666666")
    cat_font = Font(name="Calibri", bold=True, size=11, color="004D1A")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = "MKJ SUPA CUP Competition Management System - Full Audit Report"
    ws["A1"].font = title_font
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {now.strftime('%B %d, %Y at %I:%M %p')}  |  Total records: {logs.count()}"
    ws["A2"].font = sub_font

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws.merge_cells("A4:B4")
    ws["A4"] = "Category Summary"
    ws["A4"].font = cat_font
    row = 5
    for cat_name, cat_actions in ACTION_CATEGORIES.items():
        count = logs.filter(action__in=cat_actions).count()
        if count:
            ws.cell(row=row, column=1, value=cat_name).border = thin
            ws.cell(row=row, column=2, value=count).border = thin
            row += 1
    row += 1

    # ── Column headers ────────────────────────────────────────────────────
    headers = ["#", "Date & Time", "User (Email)", "User (Name)", "Role", "Category", "Action", "Description"]
    col_widths = [6, 22, 30, 25, 18, 22, 24, 60]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w

    hdr_row = row
    row += 1

    # ── Data rows ─────────────────────────────────────────────────────────
    for idx, log in enumerate(logs[:5000], start=1):
        ts = timezone.localtime(log.timestamp)
        u_email = log.user.email if log.user else "System"
        u_name = log.user.get_full_name() if log.user else "-"
        u_role = log.user.get_role_display() if log.user else "-"
        cat = ACTION_TO_CATEGORY.get(log.action, "Other")

        vals = [idx, ts.strftime("%Y-%m-%d %H:%M:%S"), u_email, u_name, u_role,
                cat, log.get_action_display(), log.description]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 8))
        row += 1

    # Auto-filter
    ws.auto_filter.ref = f"A{hdr_row}:H{row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"MKJ SUPA CUP_Audit_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@user_passes_test(superadmin_required)
def export_audit_pdf(request):
    """Export full audit report to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    logs = _filtered_audit(request)
    now = timezone.localtime(timezone.now())

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    green = colors.HexColor("#004D1A")

    title_s = ParagraphStyle("T", parent=styles["Heading1"], fontSize=16, textColor=green, spaceAfter=4)
    sub_s = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.gray, spaceAfter=12)
    cell_s = ParagraphStyle("C", parent=styles["Normal"], fontSize=7, leading=9)
    cat_s = ParagraphStyle("Cat", parent=styles["Heading3"], fontSize=10, textColor=green, spaceBefore=8, spaceAfter=4)

    elems = []

    # ── Title ─────────────────────────────────────────────────────────────
    elems.append(Paragraph("MKJ SUPA CUP Competition Management System", title_s))
    elems.append(Paragraph("Comprehensive Audit Report", ParagraphStyle(
        "S2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#333"), spaceAfter=4)))
    elems.append(Paragraph(
        f"Generated: {now.strftime('%B %d, %Y at %I:%M %p')} &bull; Total records: {logs.count()}", sub_s))
    elems.append(Spacer(1, 4 * mm))

    # ── Category summary table ────────────────────────────────────────────
    elems.append(Paragraph("Action Category Summary", cat_s))
    sum_data = [["Category", "Count"]]
    for cat_name, cat_actions in ACTION_CATEGORIES.items():
        c = logs.filter(action__in=cat_actions).count()
        if c:
            sum_data.append([cat_name, str(c)])
    if len(sum_data) > 1:
        st = Table(sum_data, colWidths=[180, 60], repeatRows=1)
        st.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), green),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F7F0")]),
        ]))
        elems.append(st)
    elems.append(Spacer(1, 6 * mm))

    # ── Full log table ────────────────────────────────────────────────────
    elems.append(Paragraph("Detailed Activity Log", cat_s))
    header = ["#", "Date & Time", "User", "Role", "Category", "Action", "Description"]
    data = [header]
    for i, log in enumerate(logs[:2000], 1):
        ts = timezone.localtime(log.timestamp)
        u = log.user.email if log.user else "System"
        r = log.user.get_role_display() if log.user else "-"
        cat = ACTION_TO_CATEGORY.get(log.action, "Other")
        desc = log.description[:100] + ("…" if len(log.description) > 100 else "")
        data.append([str(i), ts.strftime("%Y-%m-%d %H:%M"), Paragraph(u, cell_s), r,
                      cat, log.get_action_display(), Paragraph(desc, cell_s)])

    cw = [20, 58, 90, 55, 70, 75, 210]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), green),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("FONTSIZE", (0, 1), (-1, -1), 6.5),
        ("TOPPADDING", (0, 1), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F7F0")]),
    ]))
    elems.append(tbl)

    # Footer
    elems.append(Spacer(1, 8 * mm))
    elems.append(Paragraph(
        f"<i>Audit report generated by MKJ SUPA CUP CMS on {now.strftime('%d/%m/%Y %H:%M')}. "
        f"Confidential - for authorised administrators only.</i>",
        ParagraphStyle("Ft", parent=styles["Normal"], fontSize=7, textColor=colors.gray),
    ))

    doc.build(elems)
    buf.seek(0)
    fn = f"MKJ SUPA CUP_Audit_Report_{now.strftime('%Y%m%d_%H%M')}.pdf"
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp
