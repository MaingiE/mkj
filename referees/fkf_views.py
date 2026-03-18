from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import timedelta
from matches.models import Match
from .models import MatchOfficials, MatchReport, SquadEditRequest
from tournaments.models import TournamentMatchOfficials, TournamentMatch

def referees_manager_required(user):
    """Check if user is in Referees Manager group or is staff"""
    return user.groups.filter(name='Referees Manager').exists() or user.is_staff

@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
def cancel_appointment(request, match_id):
    """Cancel all officials for a match (Referees Manager only)"""
    match = get_object_or_404(Match, id=match_id)
    try:
        officials = MatchOfficials.objects.get(match=match)
    except MatchOfficials.DoesNotExist:
        messages.error(request, "No officials appointed for this match.")
        return redirect('matches:match_details', match_id=match.id)

    if request.method == 'POST':
        # Collect all appointed referees before deletion
        appointed_referees = []
        
        if officials.main_referee:
            appointed_referees.append({'referee': officials.main_referee, 'role': 'Referee'})
        if officials.assistant_1:
            appointed_referees.append({'referee': officials.assistant_1, 'role': 'Assistant Referee 1'})
        if officials.assistant_2:
            appointed_referees.append({'referee': officials.assistant_2, 'role': 'Assistant Referee 2'})
        if officials.fourth_official:
            appointed_referees.append({'referee': officials.fourth_official, 'role': 'Reserve Referee'})
        if officials.reserve_referee:
            appointed_referees.append({'referee': officials.reserve_referee, 'role': 'Reserve Referee'})
        if officials.reserve_assistant:
            appointed_referees.append({'referee': officials.reserve_assistant, 'role': 'Reserve Assistant'})
        if officials.var:
            appointed_referees.append({'referee': officials.var, 'role': 'VAR'})
        if officials.avar1:
            appointed_referees.append({'referee': officials.avar1, 'role': 'AVAR1'})
        if officials.match_commissioner:
            appointed_referees.append({'referee': officials.match_commissioner, 'role': 'Match Commissioner'})
        
        # Delete the appointments
        officials.delete()
        
        # Send notifications to all affected referees
        for ref_data in appointed_referees:
            referee = ref_data['referee']
            role = ref_data['role']
            
            # Send Email Notification
            try:
                from django.core.mail import EmailMultiAlternatives
                from django.conf import settings
                
                subject = f'Match Appointment Cancelled - {match.home_team} vs {match.away_team}'
                
                text_content = f"""
Dear {referee.full_name},

Your appointment for the following match has been CANCELLED by the Referees Manager:

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MATCH DETAILS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Match: {match.home_team} vs {match.away_team}
Date: {match.match_date.strftime('%A, %B %d, %Y')}
Time: {match.kickoff_time if match.kickoff_time else 'TBD'}
Venue: {match.venue if match.venue else 'TBA'}
Your Role: {role}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

This appointment has been removed from your schedule. You are no longer required to attend this match.

If you have any questions or concerns, please contact the Referees Manager.

Best regards,
FKF Meru County League - Referees Management
"""
                
                email = EmailMultiAlternatives(
                    subject,
                    text_content,
                    settings.DEFAULT_FROM_EMAIL,
                    [referee.email]
                )
                email.send(fail_silently=True)
                
            except Exception as e:
                # Log error but don't stop the process
                print(f"Failed to send email to {referee.email}: {str(e)}")
            
            # Send SMS Notification (if phone number available)
            if referee.phone_number:
                try:
                    sms_message = (
                        f"FKF MERU: Your appointment as {role} for "
                        f"{match.home_team} vs {match.away_team} on "
                        f"{match.match_date.strftime('%d/%m/%Y')} has been CANCELLED. "
                        f"Contact Referees Manager for details."
                    )
                    
                    # TODO: Integrate with SMS service (e.g., Africa's Talking)
                    # For now, we'll log it
                    print(f"SMS to {referee.phone_number}: {sms_message}")
                    
                    # Uncomment when SMS service is configured:
                    # from your_sms_service import send_sms
                    # send_sms(referee.phone_number, sms_message)
                    
                except Exception as e:
                    print(f"Failed to send SMS to {referee.phone_number}: {str(e)}")
        
        messages.success(
            request, 
            f"Referee appointments cancelled. All appointments for {match} have been cancelled. "
            f"Notifications sent to {len(appointed_referees)} referee(s)."
        )
        return redirect('dashboard')  # Redirect to referees manager dashboard

    context = {
        'match': match,
        'officials': officials,
    }
    return render(request, 'referees/cancel_appointment_confirm.html', context)
from django.http import JsonResponse
from django.views.decorators.http import require_GET
import json
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import permission_required
from django.contrib import messages
from django.forms import inlineformset_factory
from django.utils import timezone
from django import forms
from django.db import models
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, authenticate
from django.utils.safestring import mark_safe  
from django.db.models import Q
# Add these imports if not already present
from django.db import models
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime, timedelta
from django.utils import timezone
from matches.models import Match
from .models import (
    Referee, MatchReport, MatchOfficials, TeamOfficial, PlayingKit,
    MatchVenueDetails, StartingLineup, ReservePlayer, 
    Substitution, Caution, Expulsion, MatchGoal, RefereeAvailability,
    MatchdaySquad
)

from .forms import (
    RefereeRegistrationForm,
    RefereeProfileUpdateForm,
    MatchReportForm, 
    MatchOfficialsAppointmentForm,
    MatchOfficialsManualEntryForm,
    TeamOfficialForm,
    PlayingKitForm, 
    MatchVenueDetailsForm, 
    StartingLineupForm,
    ReservePlayerForm, 
    SubstitutionForm, 
    CautionForm,
    ExpulsionForm, 
    MatchGoalForm,
    MatchScoreForm
)

from matches.models import Match
from teams.models import Team, Player
from django.shortcuts import render
from django.contrib.auth.decorators import login_required


# ========== PUBLIC REFEREE REGISTRATION ==========
def referee_registration(request):
    """SIMPLE PUBLIC REGISTRATION - Only 4 required fields!"""
    if request.method == 'POST':
        form = RefereeRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            referee = form.save(commit=False)
            referee.status = 'pending'
            referee.is_active = False
            referee.save()
            
            messages.success(request, mark_safe(
                '<strong>Registration Successful!</strong><br><br>'
                f'Thank you, <strong>{referee.full_name}</strong>!<br>'
                f'Your FKF Number: <code>{referee.fkf_number}</code><br>'
                f'Your Email: <code>{referee.email}</code><br><br>'
                '<strong>Next Steps:</strong><br>'
                '1. Wait for admin approval<br>'
                '2. After approval, you will receive login credentials<br>'
                '3. Login and change your password<br>'
            ))
            # Redirect to a clearer next-step page for referees
            return redirect('referees:login_instructions')
    else:
        form = RefereeRegistrationForm()
    
    return render(request, 'referees/register.html', {'form': form})

# ========== LOGIN INSTRUCTIONS ==========
def login_instructions(request):
    """Display login instructions for referees"""
    context = {}
    
    if request.user.is_authenticated:
        try:
            referee = request.user.referee_profile
            context['referee'] = referee
            context['status'] = referee.status
        except AttributeError:
            pass
    
    return render(request, 'referees/login_instructions.html', context)

# ========== ADMIN: PENDING REFEREES ==========
@login_required
@user_passes_test(referees_manager_required)
def pending_referees(request):
    """View pending referee registrations (Referees Manager or Admin)"""
    pending_referees_list = Referee.objects.filter(status='pending').order_by('-date_joined')
    
    context = {
        'pending_referees': pending_referees_list,
        'approved_count': Referee.objects.filter(status='approved').count(),
        'rejected_count': Referee.objects.filter(status='rejected').count(),
    }
    
    return render(request, 'referees/admin/pending_referees.html', context)

# ========== ADMIN: APPROVE REFEREE ==========
@login_required
@user_passes_test(referees_manager_required)
def approve_referee(request, referee_id):
    """Approve a referee and create their account (Referees Manager or Admin)"""
    referee = get_object_or_404(Referee, id=referee_id)
    
    if request.method == 'POST':
        # Call the approve method which creates user account
        unique_id, default_password = referee.approve(request.user)
        
        # AUTO-ADD REFEREE TO "REFEREE" GROUP
        try:
            referee_group = Group.objects.get(name='Referee')
            referee.user.groups.add(referee_group)
        except Group.DoesNotExist:
            messages.warning(request, "Referee group not found. Please create it in Django Admin.")
        
        messages.success(request, mark_safe(
            f'<strong>Referee {referee.full_name} has been approved!</strong><br><br>'
            f'• <strong>Referee ID:</strong> <code>{unique_id}</code><br>'
            f'• Login credentials have been sent to the referee\'s email.<br>'
        ))
        return redirect('referees:pending_referees')
    
    context = {'referee': referee}
    return render(request, 'referees/admin/approve_referee.html', context)

# ========== ADMIN: REJECT REFEREE ==========
@login_required
@user_passes_test(referees_manager_required)
def reject_referee(request, referee_id):
    """Reject a referee registration (Referees Manager or Admin)"""
    referee = get_object_or_404(Referee, id=referee_id)
    
    if request.method == 'POST':
        reason = request.POST.get('rejection_reason', '')
        referee.reject(reason)
        
        messages.success(request, f'Referee {referee.full_name} has been rejected.')
        return redirect('referees:pending_referees')
    
    context = {'referee': referee}
    return render(request, 'referees/admin/reject_referee.html', context)

# ========== REFEREE DASHBOARD ==========
@login_required
def referee_dashboard(request):
    """Referee dashboard - only for approved referees"""
    # Check if user is a Referees Manager - redirect to manager dashboard
    if request.user.groups.filter(name='Referees Manager').exists():
        return redirect('referees:matches_needing_officials')
    
    # Check if user is in Referee group
    if not request.user.groups.filter(name='Referee').exists():
        messages.error(request, "Access denied. You are not registered as a referee.")
        return redirect('frontend:home')
    
    try:
        referee = request.user.referee_profile
        if not referee.can_be_appointed():
            messages.warning(request, mark_safe(
                '<strong>Your referee account is pending approval.</strong><br>'
                f'Status: <strong>{referee.get_status_display()}</strong>'
            ))
            return redirect('frontend:home')

        appointments = MatchOfficials.objects.filter(
            Q(main_referee=referee) | Q(assistant_1=referee) |
            Q(assistant_2=referee) | Q(fourth_official=referee) |
            Q(reserve_referee=referee) | Q(reserve_assistant=referee) |
            Q(var=referee) | Q(avar1=referee) | Q(match_commissioner=referee)
        ).select_related('match', 'match__home_team', 'match__away_team', 'match__zone').order_by('match__match_date')

        print(f"DEBUG: Found {appointments.count()} appointments for referee {referee.full_name}")

        upcoming_matches = []
        pending_confirmation = []
        current_matches = []
        completed_matches = []
        today = timezone.now().date()

        for appointment in appointments:
            match = appointment.match
            role = "Unknown"
            confirmed = False
            if appointment.main_referee == referee:
                role = "REFEREE"
                confirmed = appointment.main_confirmed
            elif appointment.assistant_1 == referee:
                role = "AR1"
                confirmed = appointment.ar1_confirmed
            elif appointment.assistant_2 == referee:
                role = "AR2"
                confirmed = appointment.ar2_confirmed
            elif appointment.fourth_official == referee:
                role = "RESERVE"
                confirmed = appointment.fourth_confirmed
            elif appointment.reserve_referee == referee:
                role = "RESERVE"
                confirmed = appointment.reserve_confirmed
            elif appointment.var == referee:
                role = "VAR"
                confirmed = appointment.var_confirmed
            elif appointment.avar1 == referee:
                role = "AVAR1"
                confirmed = False
            elif appointment.match_commissioner == referee:
                role = "COMMISSIONER"
                confirmed = appointment.commissioner_confirmed

            def get_official_display_name(official):
                if not official:
                    return None
                if hasattr(official, 'user') and official.user:
                    return official.user.get_full_name() or f"{official.first_name} {official.last_name}"
                return f"{official.first_name} {official.last_name}"

            officials_data = {
                'main_referee': get_official_display_name(appointment.main_referee),
                'assistant_1': get_official_display_name(appointment.assistant_1),
                'assistant_2': get_official_display_name(appointment.assistant_2),
                'reserve_referee': get_official_display_name(appointment.fourth_official or appointment.reserve_referee),
                'var': get_official_display_name(appointment.var),
                'avar1': get_official_display_name(appointment.avar1),
                'match_commissioner': get_official_display_name(appointment.match_commissioner),
            }

            match_info = {
                'match': match,
                'role': role,
                'confirmed': confirmed,
                'appointment': appointment,
                'officials_data': officials_data,
                'match_date': match.match_date.date() if hasattr(match.match_date, 'date') else match.match_date,
                'can_confirm': match.match_date.date() >= today and not confirmed if hasattr(match.match_date, 'date') else match.match_date >= today and not confirmed
            }

            match_date = match.match_date.date() if hasattr(match.match_date, 'date') else match.match_date
            if match_date == today:
                if not confirmed:
                    current_matches.append(match_info)
                    pending_confirmation.append(match_info)
                else:
                    upcoming_matches.append(match_info)
            elif match_date > today:
                if not confirmed:
                    pending_confirmation.append(match_info)
                else:
                    upcoming_matches.append(match_info)
            else:
                completed_matches.append(match_info)

        print(f"DEBUG: Pending: {len(pending_confirmation)}, Current: {len(current_matches)}, Upcoming: {len(upcoming_matches)}, Completed: {len(completed_matches)}")

        pending_reports = MatchReport.objects.filter(
            referee=referee,
            status='draft'
        ).order_by('-created_at')[:5]
        submitted_reports = MatchReport.objects.filter(
            referee=referee,
            status='submitted'
        )
        draft_reports = MatchReport.objects.filter(
            referee=referee,
            status='draft'
        )
        is_suspended = referee.status == 'suspended'
        suspension_reason = referee.suspension_reason if is_suspended else None
        is_manager = request.user.groups.filter(name='Referees Manager').exists()
        
        # Get pending squad edit requests for matches where this referee is appointed
        match_ids = appointments.values_list('match_id', flat=True)
        pending_edit_requests = SquadEditRequest.objects.filter(
            squad__match_id__in=match_ids,
            status='pending'
        ).select_related('squad', 'squad__match', 'squad__team')

        # ── Tournament appointments ──────────────────────────────────────
        tournament_appointments = TournamentMatchOfficials.objects.filter(
            Q(main_referee=referee) | Q(assistant_1=referee) |
            Q(assistant_2=referee) | Q(fourth_official=referee) |
            Q(match_commissioner=referee)
        ).select_related(
            'match', 'match__tournament',
            'match__home_team__team', 'match__home_team__external_team',
            'match__away_team__team', 'match__away_team__external_team',
        ).order_by('match__match_date')

        tournament_upcoming = []
        tournament_completed = []
        for t_appt in tournament_appointments:
            t_match = t_appt.match
            t_role = 'Unknown'
            if t_appt.main_referee == referee:
                t_role = 'REFEREE'
            elif t_appt.assistant_1 == referee:
                t_role = 'AR1'
            elif t_appt.assistant_2 == referee:
                t_role = 'AR2'
            elif t_appt.fourth_official == referee:
                t_role = 'RESERVE'
            elif t_appt.match_commissioner == referee:
                t_role = 'COMMISSIONER'

            def _get_name(official):
                if not official:
                    return None
                if hasattr(official, 'user') and official.user:
                    return official.user.get_full_name() or f"{official.first_name} {official.last_name}"
                return f"{official.first_name} {official.last_name}"

            t_officials_data = {
                'main_referee': _get_name(t_appt.main_referee),
                'assistant_1': _get_name(t_appt.assistant_1),
                'assistant_2': _get_name(t_appt.assistant_2),
                'fourth_official': _get_name(t_appt.fourth_official),
                'match_commissioner': _get_name(t_appt.match_commissioner),
            }

            t_match_date = t_match.match_date.date() if t_match.match_date and hasattr(t_match.match_date, 'date') else t_match.match_date

            t_info = {
                'match': t_match,
                'role': t_role,
                'appointment': t_appt,
                'officials_data': t_officials_data,
                'match_date': t_match_date,
                'home_name': t_match.home_team.display_name if t_match.home_team else 'TBD',
                'away_name': t_match.away_team.display_name if t_match.away_team else 'TBD',
                'tournament_name': t_match.tournament.name,
                'tournament_slug': t_match.tournament.slug,
            }

            if t_match_date and t_match_date >= today:
                tournament_upcoming.append(t_info)
            else:
                tournament_completed.append(t_info)

        context = {
            'referee': referee,
            'current_matches': current_matches,
            'upcoming_matches': upcoming_matches,
            'pending_confirmation': pending_confirmation,
            'completed_matches': completed_matches,
            'pending_reports': pending_reports,
            'submitted_reports': submitted_reports,
            'draft_reports': draft_reports,
            'pending_edit_requests': pending_edit_requests,
            'can_confirm': request.user.has_perm('referees.confirm_appointment'),
            'can_submit_report': request.user.has_perm('referees.submit_match_report'),
            'is_suspended': is_suspended,
            'suspension_reason': suspension_reason,
            'total_appointments': appointments.count(),
            'is_manager': is_manager,
            # Tournament
            'tournament_upcoming': tournament_upcoming,
            'tournament_completed': tournament_completed,
        }
        return render(request, 'referees/dashboard.html', context)
    
    except AttributeError as e:
        print(f"DEBUG: AttributeError in referee_dashboard: {str(e)}")
        messages.error(request, "You need to register as a referee first.")
        return redirect('referees:referee_registration')

# ========== REFEREE PROFILE ==========
@login_required
def referee_profile(request):
    """View and update referee profile"""
    try:
        referee = request.user.referee_profile
        
        if request.method == 'POST':
            form = RefereeProfileUpdateForm(request.POST, request.FILES, instance=referee)
            if form.is_valid():
                form.save()
                messages.success(request, 'Profile updated successfully!')
                return redirect('referees:referee_profile')
        else:
            form = RefereeProfileUpdateForm(instance=referee)
        
        context = {
            'referee': referee,
            'form': form,
        }
        return render(request, 'referees/profile.html', context)
        
    except AttributeError:
        messages.error(request, "You don't have a referee profile.")
        return redirect('referees:referee_registration')

# ========== REFEREE: CONFIRM APPOINTMENT ==========
@login_required
@permission_required('referees.confirm_appointment', raise_exception=True)
def confirm_appointment(request, match_id):
    """Referee confirms their appointment to a match"""
    try:
        referee = request.user.referee_profile
        
        if not referee.can_be_appointed():
            messages.error(request, "Your account is not approved or active.")
            return redirect('referees:referee_dashboard')
            
    except AttributeError:
        messages.error(request, "You need a referee profile to confirm appointments.")
        return redirect('referees:referee_registration')
    
    match = get_object_or_404(Match, id=match_id)
    
    # Check if match has officials
    if not hasattr(match, 'officials'):
        messages.error(request, "No officials appointed for this match.")
        return redirect('referees:referee_dashboard')
    
    officials = match.officials
    
    # Determine referee's role in this match
    role = None
    is_appointed = False
    
    if officials.main_referee == referee:
        role = 'REFEREE'
        is_appointed = True
        already_confirmed = officials.main_confirmed
    elif officials.assistant_1 == referee:
        role = 'AR1'
        is_appointed = True
        already_confirmed = officials.ar1_confirmed
    elif officials.assistant_2 == referee:
        role = 'AR2'
        is_appointed = True
        already_confirmed = officials.ar2_confirmed
    elif officials.fourth_official == referee:
        role = 'FOURTH'
        is_appointed = True
        already_confirmed = officials.fourth_confirmed
    elif officials.reserve_referee == referee:
        role = 'RESERVE'
        is_appointed = True
        already_confirmed = officials.reserve_confirmed
    elif officials.var == referee:
        role = 'VAR'
        is_appointed = True
        already_confirmed = officials.var_confirmed
    elif officials.avar1 == referee:
        role = 'AVAR1'
        is_appointed = True
        already_confirmed = False
    elif officials.match_commissioner == referee:
        role = 'COMMISSIONER'
        is_appointed = True
        already_confirmed = officials.commissioner_confirmed
    
    if not is_appointed:
        messages.error(request, "You are not appointed to this match.")
        return redirect('referees:referee_dashboard')
    
    if request.method == 'POST':
        # Update confirmation based on role
        if role == 'REFEREE':
            officials.main_confirmed = True
            officials.main_confirmed_at = timezone.now()
            messages.success(request, "Referee appointment confirmed!")
        elif role == 'AR1':
            officials.ar1_confirmed = True
            officials.ar1_confirmed_at = timezone.now()
            messages.success(request, "Assistant Referee 1 appointment confirmed!")
        elif role == 'AR2':
            officials.ar2_confirmed = True
            officials.ar2_confirmed_at = timezone.now()
            messages.success(request, "Assistant Referee 2 appointment confirmed!")
        elif role == 'FOURTH':
            officials.fourth_confirmed = True
            messages.success(request, "Fourth Official appointment confirmed!")
        elif role == 'RESERVE':
            officials.reserve_confirmed = True
            messages.success(request, "Reserve Referee appointment confirmed!")
        elif role == 'VAR':
            officials.var_confirmed = True
            messages.success(request, "VAR appointment confirmed!")
        elif role == 'COMMISSIONER':
            officials.commissioner_confirmed = True
            officials.commissioner_confirmed_at = timezone.now()
            messages.success(request, "Match Commissioner appointment confirmed!")
        
        # Update status if all required officials confirmed
        if officials.all_required_confirmed:
            officials.status = 'CONFIRMED'
        
        officials.save()
        return redirect('referees:referee_dashboard')
    
    context = {
        'match': match,
        'officials': officials,
        'role': role,
        'already_confirmed': already_confirmed,
        'role_display': dict(officials.OFFICIAL_ROLES).get(role, role),
        'match_date': match.match_date.date(),
        'today': timezone.now().date(),
    }
    
    return render(request, 'referees/confirm_appointment.html', context)


# ========== REFEREE: DECLINE APPOINTMENT ==========
@login_required
@permission_required('referees.confirm_appointment', raise_exception=True)
def decline_appointment(request, match_id):
    """Referee declines their appointment with a reason"""
    try:
        referee = request.user.referee_profile
        
        if not referee.can_be_appointed():
            messages.error(request, "Your account is not approved or active.")
            return redirect('referees:referee_dashboard')
            
    except AttributeError:
        messages.error(request, "You need a referee profile to reject appointments.")
        return redirect('referees:referee_registration')
    
    match = get_object_or_404(Match, id=match_id)
    
    # Check if match has officials
    if not hasattr(match, 'officials'):
        messages.error(request, "No officials appointed for this match.")
        return redirect('referees:referee_dashboard')
    
    officials = match.officials
    
    # Determine referee's role in this match
    role = None
    is_appointed = False
    
    if officials.main_referee == referee:
        role = 'REFEREE'
        is_appointed = True
        already_rejected = officials.main_rejected
    elif officials.assistant_1 == referee:
        role = 'AR1'
        is_appointed = True
        already_rejected = officials.ar1_rejected
    elif officials.assistant_2 == referee:
        role = 'AR2'
        is_appointed = True
        already_rejected = officials.ar2_rejected
    elif officials.fourth_official == referee:
        role = 'FOURTH'
        is_appointed = True
        already_rejected = officials.fourth_rejected
    elif officials.reserve_referee == referee:
        role = 'RESERVE'
        is_appointed = True
        already_rejected = officials.reserve_rejected
    elif officials.var == referee:
        role = 'VAR'
        is_appointed = True
        already_rejected = officials.var_rejected
    elif officials.match_commissioner == referee:
        role = 'COMMISSIONER'
        is_appointed = True
        already_rejected = officials.commissioner_rejected
    
    if not is_appointed:
        messages.error(request, "You are not appointed to this match.")
        return redirect('referees:referee_dashboard')
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        
        if not reason:
            messages.error(request, "Please provide a reason for declining this appointment.")
            return render(request, 'referees/decline_appointment.html', {
                'match': match,
                'officials': officials,
                'role': role,
                'role_display': dict(officials.OFFICIAL_ROLES).get(role, role),
            })
        
        # Update rejection based on role
        if role == 'REFEREE':
            officials.main_rejected = True
            officials.main_rejection_reason = reason
            officials.main_rejected_at = timezone.now()
            messages.success(request, "Referee appointment declined. Referees Manager will be notified.")
        elif role == 'AR1':
            officials.ar1_rejected = True
            officials.ar1_rejection_reason = reason
            officials.ar1_rejected_at = timezone.now()
            messages.success(request, "Assistant Referee 1 appointment declined. Referees Manager will be notified.")
        elif role == 'AR2':
            officials.ar2_rejected = True
            officials.ar2_rejection_reason = reason
            officials.ar2_rejected_at = timezone.now()
            messages.success(request, "Assistant Referee 2 appointment declined. Referees Manager will be notified.")
        elif role == 'FOURTH':
            officials.fourth_rejected = True
            officials.fourth_rejection_reason = reason
            messages.success(request, "Fourth Official appointment declined. Referees Manager will be notified.")
        elif role == 'RESERVE':
            officials.reserve_rejected = True
            officials.reserve_rejection_reason = reason
            messages.success(request, "Reserve Referee appointment declined. Referees Manager will be notified.")
        elif role == 'VAR':
            officials.var_rejected = True
            officials.var_rejection_reason = reason
            messages.success(request, "VAR appointment declined. Referees Manager will be notified.")
        elif role == 'COMMISSIONER':
            officials.commissioner_rejected = True
            officials.commissioner_rejection_reason = reason
            messages.success(request, "Match Commissioner appointment declined. Referees Manager will be notified.")
        
        # Update status to REJECTED if any official rejected
        if any([officials.main_rejected, officials.ar1_rejected, officials.ar2_rejected,
                officials.fourth_rejected, officials.reserve_rejected, officials.var_rejected,
                officials.commissioner_rejected]):
            officials.status = 'REJECTED'
        
        officials.save()
        return redirect('referees:referee_dashboard')
    
    context = {
        'match': match,
        'officials': officials,
        'role': role,
        'already_rejected': already_rejected,
        'role_display': dict(officials.OFFICIAL_ROLES).get(role, role),
        'match_date': match.match_date.date(),
        'today': timezone.now().date(),
    }
    
    return render(request, 'referees/decline_appointment.html', context)


# ========== APPOINT MATCH OFFICIALS ==========
@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
def appoint_match_officials(request, match_id):
    """Appoint officials to a match - Referees Manager only"""
    match = get_object_or_404(Match, id=match_id)
    
    # Check 4-day rule
    match_date = match.match_date.date() if hasattr(match.match_date, 'date') else match.match_date
    days_until_match = (match_date - timezone.now().date()).days
    
    if days_until_match > 4:
        messages.error(request, 
            f"Cannot appoint officials yet. Match is in {days_until_match} days. "
            "Officials can only be appointed within 4 days of match."
        )
        return redirect('matches:match_detail', match_id=match.id)
    
    # Get or create MatchOfficials instance
    match_officials, created = MatchOfficials.objects.get_or_create(match=match)
    
    if request.method == 'POST':
        appointment_form = MatchOfficialsAppointmentForm(request.POST, instance=match_officials)
        
        if appointment_form.is_valid():
            officials = appointment_form.save(commit=False)
            officials.status = 'APPOINTED'
            officials.appointment_made_by = request.user
            try:
                officials.save()
                # Reset confirmations when new officials are appointed
                officials.main_confirmed = False
                officials.ar1_confirmed = False
                officials.ar2_confirmed = False
                officials.reserve_confirmed = False
                officials.var_confirmed = False
                officials.save()
                messages.success(request, f'Match officials appointed for {match}!')
                return redirect('referees:appoint_match_officials', match_id=match.id)
            except ValidationError as ve:
                error_msg = '; '.join([str(msg) for msgs in ve.message_dict.values() for msg in msgs])
                messages.error(request, error_msg)
            except Exception as e:
                messages.error(request, f'Error appointing officials: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        appointment_form = MatchOfficialsAppointmentForm(instance=match_officials)
    
    # Get available referees
    available_referees = Referee.objects.filter(
        status='approved',
        is_active=True
    )
    
    # Check availability for match date
    for referee in available_referees:
        referee.is_available = RefereeAvailability.objects.filter(
            referee=referee,
            date=match_date,
            is_available=False
        ).exists()
    
    context = {
        'match': match,
        'form': appointment_form,
        'match_officials': match_officials,
        'available_referees': available_referees,
        'days_until_match': days_until_match,
        'can_replace': request.user.has_perm('referees.replace_referee'),
        'OFFICIAL_ROLES': MatchOfficials.OFFICIAL_ROLES,
    }
    
    return render(request, 'referees/appoint_officials.html', context)

# ========== REPLACE APPOINTED REFEREE ==========
@login_required
@permission_required('referees.replace_referee', raise_exception=True)
def replace_referee(request, match_id, role):
    """Replace an appointed referee (Referees Manager only)"""
    match = get_object_or_404(Match, id=match_id)
    
    if not hasattr(match, 'officials'):
        messages.error(request, "No officials appointed for this match.")
        return redirect('matches:match_details', match_id=match.id)
    
    officials = match.officials
    
    # Map role parameter to model field
    role_mapping = {
        'REFEREE': {'field': 'main_referee', 'name': 'Referee'},
        'AR1': {'field': 'assistant_1', 'name': 'Assistant Referee 1'},
        'AR2': {'field': 'assistant_2', 'name': 'Assistant Referee 2'},
        'RESERVE': {'field': 'reserve_referee', 'name': 'Reserve Referee'},
        'RESERVE_AR': {'field': 'reserve_assistant', 'name': 'Reserve Assistant Referee'},
        'VAR': {'field': 'var', 'name': 'Video Assistant Referee'},
        'AVAR1': {'field': 'avar1', 'name': 'Assistant VAR 1'},
        'FOURTH': {'field': 'fourth_official', 'name': 'Reserve Referee'},
        'COMMISSIONER': {'field': 'match_commissioner', 'name': 'Match Commissioner'},
    }
    
    if role not in role_mapping:
        messages.error(request, "Invalid role specified.")
        return redirect('referees:appoint_match_officials', match_id=match.id)
    
    role_info = role_mapping[role]
    current_referee = getattr(officials, role_info['field'])
    
    if request.method == 'POST':
        new_referee_id = request.POST.get('new_referee')
        
        if not new_referee_id:
            messages.error(request, "Please select a referee.")
            return redirect('referees:replace_referee', match_id=match.id, role=role)
        
        new_referee = get_object_or_404(Referee, id=new_referee_id)
        
        # Check if new referee can be appointed
        if not new_referee.can_be_appointed():
            messages.error(request, 
                f"{new_referee.full_name} cannot be appointed. "
                f"Status: {new_referee.get_status_display()}"
            )
            return redirect('referees:replace_referee', match_id=match.id, role=role)
        
        # Check availability
        unavailable = RefereeAvailability.objects.filter(
            referee=new_referee,
            date=match.match_date.date(),
            is_available=False
        ).first()
        
        if unavailable:
            messages.error(request,
                f"{new_referee.full_name} is unavailable on {match.match_date.date()}. "
                f"Reason: {unavailable.reason}"
            )
            return redirect('referees:replace_referee', match_id=match.id, role=role)
        
        # Replace the referee
        setattr(officials, role_info['field'], new_referee)
        
        # Reset confirmation for this role
        if role == 'REFEREE':
            officials.main_confirmed = False
            officials.main_confirmed_at = None
        elif role == 'AR1':
            officials.ar1_confirmed = False
            officials.ar1_confirmed_at = None
        elif role == 'AR2':
            officials.ar2_confirmed = False
            officials.ar2_confirmed_at = None
        elif role == 'RESERVE':
            officials.reserve_confirmed = False
        elif role == 'VAR':
            officials.var_confirmed = False
        elif role == 'FOURTH':
            officials.fourth_confirmed = False
        
        officials.save()
        
        messages.success(request,
            f"{role_info['name']} replaced successfully! "
            f"Old: {current_referee.full_name if current_referee else 'None'} → "
            f"New: {new_referee.full_name}"
        )
        
        return redirect('referees:appoint_match_officials', match_id=match.id)
    
    # Get available referees (excluding current one)
    available_referees = Referee.objects.filter(
        status='approved',
        is_active=True
    ).exclude(
        id=current_referee.id if current_referee else None
    )
    
    context = {
        'match': match,
        'officials': officials,
        'role': role,
        'role_name': role_info['name'],
        'current_referee': current_referee,
        'available_referees': available_referees,
        'match_date': match.match_date.date(),
    }
    
    return render(request, 'referees/replace_referee.html', context)

# ========== MATCHES NEEDING OFFICIALS ==========
@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
def matches_needing_officials(request):
    """Show matches that need officials appointed (Referees Manager)"""
    # Get matches within 4 days that don't have officials or are pending
    four_days_from_now = timezone.now() + timezone.timedelta(days=4)
    
    matches = Match.objects.filter(
        match_date__lte=four_days_from_now,
        match_date__gte=timezone.now(),
        status='scheduled'
    ).select_related('home_team', 'away_team', 'zone').prefetch_related('officials').order_by('round_number', 'match_date')
    
    # Categorize matches by appointment status
    needs_officials = []
    appointed_matches = []
    pending_confirmation = []
    confirmed_matches = []
    declined_appointments = []
    
    for match in matches:
        if not hasattr(match, 'officials') or match.officials is None:
            # Only add to needs_officials if no appointment exists
            match.has_declined = False
            needs_officials.append(match)
        else:
            # Check if any officials have declined
            has_declined = any([
                match.officials.main_rejected,
                match.officials.ar1_rejected,
                match.officials.ar2_rejected,
                match.officials.fourth_rejected,
                match.officials.reserve_rejected,
                match.officials.var_rejected,
                match.officials.commissioner_rejected
            ])
            
            if has_declined:
                # Add declined reasons to match object for display
                match.has_declined = True
                match.declined_roles = []
                if match.officials.main_rejected:
                    match.declined_roles.append({
                        'role': 'Main Referee',
                        'referee': match.officials.main_referee,
                        'reason': match.officials.main_rejection_reason,
                        'declined_at': match.officials.main_rejected_at
                    })
                if match.officials.ar1_rejected:
                    match.declined_roles.append({
                        'role': 'Assistant Referee 1',
                        'referee': match.officials.assistant_1,
                        'reason': match.officials.ar1_rejection_reason,
                        'declined_at': match.officials.ar1_rejected_at
                    })
                if match.officials.ar2_rejected:
                    match.declined_roles.append({
                        'role': 'Assistant Referee 2',
                        'referee': match.officials.assistant_2,
                        'reason': match.officials.ar2_rejection_reason,
                        'declined_at': match.officials.ar2_rejected_at
                    })
                if match.officials.fourth_rejected:
                    match.declined_roles.append({
                        'role': 'Fourth Official',
                        'referee': match.officials.fourth_official,
                        'reason': match.officials.fourth_rejection_reason,
                        'declined_at': getattr(match.officials, 'fourth_rejected_at', None)
                    })
                if match.officials.reserve_rejected:
                    match.declined_roles.append({
                        'role': 'Reserve Referee',
                        'referee': match.officials.reserve_referee,
                        'reason': match.officials.reserve_rejection_reason,
                        'declined_at': getattr(match.officials, 'reserve_rejected_at', None)
                    })
                if match.officials.var_rejected:
                    match.declined_roles.append({
                        'role': 'VAR',
                        'referee': match.officials.var,
                        'reason': match.officials.var_rejection_reason,
                        'declined_at': getattr(match.officials, 'var_rejected_at', None)
                    })
                if match.officials.commissioner_rejected:
                    match.declined_roles.append({
                        'role': 'Match Commissioner',
                        'referee': match.officials.match_commissioner,
                        'reason': match.officials.commissioner_rejection_reason,
                        'declined_at': getattr(match.officials, 'commissioner_rejected_at', None)
                    })
                declined_appointments.append(match)
                # DO NOT add to needs_officials - keep them separate
            elif match.officials.status == 'APPOINTED':
                appointed_matches.append(match)
                pending_confirmation.append(match)
            elif match.officials.status == 'CONFIRMED':
                appointed_matches.append(match)
                confirmed_matches.append(match)
    
    # ── Tournament matches needing officials ─────────────────────────────
    t_needs_officials = TournamentMatch.objects.filter(
        Q(officials__isnull=True) | Q(officials__status='PENDING'),
        status='scheduled',
        match_date__gte=timezone.now(),
    ).select_related(
        'tournament',
        'home_team__team', 'home_team__external_team',
        'away_team__team', 'away_team__external_team',
    ).order_by('match_date')

    t_appointed = TournamentMatch.objects.filter(
        officials__status__in=['APPOINTED', 'CONFIRMED'],
        status='scheduled',
        match_date__gte=timezone.now(),
    ).select_related(
        'tournament', 'officials',
        'home_team__team', 'home_team__external_team',
        'away_team__team', 'away_team__external_team',
    ).order_by('match_date')

    context = {
        'needs_officials': needs_officials,
        'appointed_matches': appointed_matches,
        'pending_confirmation': pending_confirmation,
        'confirmed_matches': confirmed_matches,
        'declined_appointments': declined_appointments,
        'declined_count': len(declined_appointments),
        'today': timezone.now().date(),
        'four_days_later': four_days_from_now.date(),
        'OFFICIAL_ROLES': MatchOfficials.OFFICIAL_ROLES,
        # Tournament
        't_needs_officials': t_needs_officials,
        't_appointed': t_appointed,
    }
    
    return render(request, 'referees/matches_needing_officials.html', context)

# ========== REFEREE AVAILABILITY ==========
@login_required
def referee_availability(request):
    """Referee sets their availability"""
    try:
        referee = request.user.referee_profile
        
        if not referee.can_be_appointed():
            messages.warning(request, 'You cannot set availability until approved.')
            return redirect('frontend:home')
    
    except AttributeError:
        messages.error(request, 'You are not registered as a referee.')
        return redirect('referees:referee_registration')
    
    # Get existing availabilities
    availabilities = RefereeAvailability.objects.filter(
        referee=referee,
        date__gte=timezone.now().date()
    ).order_by('date')
    
    if request.method == 'POST':
        date_str = request.POST.get('date')
        is_available = request.POST.get('is_available') == 'true'
        reason = request.POST.get('reason', '')
        
        try:
            date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Update or create availability
            availability, created = RefereeAvailability.objects.update_or_create(
                referee=referee,
                date=date,
                defaults={
                    'is_available': is_available,
                    'reason': reason if not is_available else ''
                }
            )
            
            status = "Available" if is_available else "Unavailable"
            messages.success(request, f'Availability set to {status} for {date}.')
            
        except ValueError:
            messages.error(request, 'Invalid date format.')
        
        return redirect('referees:referee_availability')
    
    context = {
        'referee': referee,
        'availabilities': availabilities,
        'today': timezone.now().date(),
        'next_30_days': [timezone.now().date() + timezone.timedelta(days=i) for i in range(30)],
    }
    
    return render(request, 'referees/availability.html', context)

# ========== ADMIN: VIEW ALL REFEREES ==========
@login_required
@user_passes_test(referees_manager_required)
def all_referees(request):
    """View all referees with filtering"""
    status_filter = request.GET.get('status', '')
    active_filter = request.GET.get('active', '')
    specialization_filter = request.GET.get('specialization', '')
    
    referees = Referee.objects.all()
    
    if status_filter:
        referees = referees.filter(status=status_filter)
    
    if active_filter == 'active':
        referees = referees.filter(is_active=True)
    elif active_filter == 'inactive':
        referees = referees.filter(is_active=False)
    
    if specialization_filter:
        referees = referees.filter(specialization=specialization_filter)
    
    referees = referees.order_by('-date_joined')
    
    context = {
        'referees': referees,
        'total_count': Referee.objects.count(),
        'approved_count': Referee.objects.filter(status='approved').count(),
        'pending_count': Referee.objects.filter(status='pending').count(),
        'rejected_count': Referee.objects.filter(status='rejected').count(),
        'suspended_count': Referee.objects.filter(status='suspended').count(),
    }
    
    return render(request, 'referees/admin/all_referees.html', context)

# ========== SUSPEND REFEREE (ADMIN) ==========
@login_required
@user_passes_test(referees_manager_required)
def suspend_referee(request, referee_id):
    """Suspend an active referee (admin only)"""
    referee = get_object_or_404(Referee, id=referee_id)
    
    if request.method == 'POST':
        reason = request.POST.get('suspension_reason', '')
        referee.suspend(reason)
        
        messages.success(request, f'Referee {referee.full_name} has been suspended.')
        return redirect('referees:all_referees')
    
    context = {'referee': referee}
    return render(request, 'referees/admin/suspend_referee.html', context)

# ========== REACTIVATE REFEREE (ADMIN) ==========
@login_required
@user_passes_test(referees_manager_required)
def reactivate_referee(request, referee_id):
    """Reactivate a suspended referee (admin only)"""
    referee = get_object_or_404(Referee, id=referee_id)
    
    if request.method == 'POST':
        referee.reactivate()
        
        messages.success(request, f'Referee {referee.full_name} has been reactivated.')
        return redirect('referees:all_referees')
    
    context = {'referee': referee}
    return render(request, 'referees/admin/reactivate_referee.html', context)

# ========== ADMIN REFEREE DASHBOARD ==========
@login_required
@user_passes_test(referees_manager_required)
def admin_referee_dashboard(request):
    """Admin dashboard for referee management"""
    context = {
        'pending_referees': Referee.objects.filter(status='pending').order_by('-date_joined')[:10],
        'pending_count': Referee.objects.filter(status='pending').count(),
        'approved_count': Referee.objects.filter(status='approved').count(),
        'rejected_count': Referee.objects.filter(status='rejected').count(),
        'total_count': Referee.objects.count(),
        'recent_reports': MatchReport.objects.all().order_by('-created_at')[:10],
        'pending_report_count': MatchReport.objects.filter(status='submitted').count(),
        'approved_reports_count': MatchReport.objects.filter(status='approved').count(),
        'pending_report_list': MatchReport.objects.filter(status='submitted').select_related('match', 'referee').order_by('-submitted_at')[:5],
        'matches_needing_officials': Match.objects.filter(
            match_date__gte=timezone.now(),
            match_date__lte=timezone.now() + timezone.timedelta(days=4)
        ).filter(
            models.Q(officials__isnull=True) | 
            models.Q(officials__status='PENDING')
        ).count(),
    }
    
    return render(request, 'referees/admin_dashboard.html', context)
# ========== API VIEWS FOR REFEREES MANAGER DASHBOARD ==========

@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
@require_GET
def api_urgent_matches(request):
    """API: Get urgent matches needing officials (without appointments)"""
    from django.utils.dateformat import format
    from datetime import datetime
    
    today = timezone.localdate()
    four_days_from_now = today + timedelta(days=4)
    
    # Get matches that don't have ANY officials appointed yet
    from matches.models import Match
    all_matches = Match.objects.filter(
        match_date__date__lte=four_days_from_now,
        match_date__date__gte=today,
        status='scheduled'
    ).prefetch_related('officials').order_by('match_date')
    
    # Filter to only include matches without appointments
    matches_needing_officials = []
    for match in all_matches:
        if not hasattr(match, 'officials') or match.officials is None:
            matches_needing_officials.append(match)
    
    # Limit to 10
    matches = matches_needing_officials[:10]
    
    matches_data = []
    for match in matches:
        match_date = match.match_date.date() if hasattr(match.match_date, 'date') else match.match_date
        days_until = (match_date - today).days
        
        matches_data.append({
            'id': match.id,
            'date': match.match_date.strftime('%d/%m/%Y'),
            'time': match.kickoff_time or 'TBD',
            'home_team': match.home_team.team_name,
            'away_team': match.away_team.team_name,
            'venue': match.venue,
            'zone': match.zone.name,
            'days_until_match': days_until,
            'status': 'Needs Officials',
        })
    
    return JsonResponse({'matches': matches_data})

@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
@require_GET
def api_appointed_matches(request):
    """API: Get all matches with officials appointed"""
    from django.utils.dateformat import format
    from datetime import datetime
    
    today = timezone.localdate()
    four_days_from_now = today + timedelta(days=4)
    
    # Get all matches within 4 days that have appointments
    from matches.models import Match
    all_matches = Match.objects.filter(
        match_date__date__lte=four_days_from_now,
        match_date__date__gte=today,
        status='scheduled'
    ).prefetch_related('officials').order_by('match_date')
    
    # Filter to only include matches WITH appointments
    appointed_matches = []
    for match in all_matches:
        if hasattr(match, 'officials') and match.officials is not None:
            appointed_matches.append(match)
    
    matches_data = []
    for match in appointed_matches[:15]:  # Limit to 15 most recent
        officials = match.officials
        
        matches_data.append({
            'id': match.id,
            'date': match.match_date.strftime('%d/%m/%Y'),
            'time': match.kickoff_time or 'TBD',
            'home_team': match.home_team.team_name,
            'away_team': match.away_team.team_name,
            'zone': match.zone.name,
            'round': match.round_number,
            'referee': officials.main_referee.full_name if officials.main_referee else None,
            'ar1': officials.assistant_1.full_name if officials.assistant_1 else None,
            'ar2': officials.assistant_2.full_name if officials.assistant_2 else None,
            'status': officials.status,
            'status_display': 'Confirmed' if officials.status == 'CONFIRMED' else 'Pending Confirmation',
        })
    
    return JsonResponse({'matches': matches_data})

@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
@require_GET
def api_recent_appointments(request):
    """API: Get recent appointments made by current user"""
    from django.utils.timesince import timesince
    
    recent_appointments = MatchOfficials.objects.filter(
        appointment_made_by=request.user
    ).select_related('match').order_by('-appointment_made_at')[:10]
    
    appointments_data = []
    for appointment in recent_appointments:
        match = appointment.match
        
        appointments_data.append({
            'date': match.match_date.strftime('%d/%m'),
            'match': f"{match.home_team} vs {match.away_team}",
            'referee': appointment.main_referee.full_name if appointment.main_referee else 'Not set',
            'ar1': appointment.assistant_1.full_name if appointment.assistant_1 else 'Not set',
            'ar2': appointment.assistant_2.full_name if appointment.assistant_2 else 'Not set',
            'status': appointment.get_status_display(),
            'status_class': 'bg-success' if appointment.status == 'CONFIRMED' else 'bg-warning',
            'appointed_by': request.user.get_full_name() or request.user.username,
        })
    
    return JsonResponse({'appointments': appointments_data})

@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
@require_GET
def api_available_referees_today(request):
    """API: Get referees available today"""
    today = timezone.now().date()
    
    # Get referees who are not marked unavailable today
    available_referees = Referee.objects.filter(
        status='approved',
        is_active=True
    ).exclude(
        availabilities__date=today,
        availabilities__is_available=False
    )[:10]
    
    referees_data = []
    for referee in available_referees:
        referees_data.append({
            'id': referee.id,
            'name': referee.full_name,
            'initials': f"{referee.first_name[0]}{referee.last_name[0]}",
            'level': referee.get_level_display() or 'Not set',
            'county': referee.county or 'Not set',
        })
    
    return JsonResponse({'referees': referees_data})

@login_required
@permission_required('referees.appoint_referees', raise_exception=True)
@require_GET
def api_manager_stats(request):
    """API: Get dashboard statistics for Referees Manager"""
    today = timezone.now().date()
    tomorrow = today + timedelta(days=1)
    next_week = today + timedelta(days=7)
    now = timezone.now()
    
    # Matches needing officials (within 4 days)
    four_days_from_now = now + timedelta(days=4)
    needs_officials = Match.objects.filter(
        match_date__lte=four_days_from_now,
        match_date__gte=now,
        status='scheduled'
    ).exclude(officials__status='CONFIRMED').count()
    
    # Matches pending confirmation
    pending_confirmation = Match.objects.filter(
        officials__status='APPOINTED',
        match_date__gte=now
    ).count()
    
    # Available referees
    available_referees = Referee.objects.filter(
        status='approved',
        is_active=True
    ).exclude(
        availabilities__date=today,
        availabilities__is_available=False
    ).count()
    
    # Today's matches
    today_matches = Match.objects.filter(
        match_date__date=today,
        status='scheduled'
    ).count()
    
    # Deadlines
    appointed_matches = Match.objects.filter(
        officials__status='APPOINTED',
        match_date__gte=now
    ).select_related('officials')

    deadline_today = 0
    deadline_tomorrow = 0
    deadline_week = 0
    for match in appointed_matches:
        if hasattr(match, 'officials') and match.officials:
            deadline = match.officials.confirmation_deadline
            if deadline == today:
                deadline_today += 1
            if deadline == tomorrow:
                deadline_tomorrow += 1
            if today <= deadline <= next_week:
                deadline_week += 1
    
    return JsonResponse({
        'needs_officials': needs_officials,
        'pending_confirmation': pending_confirmation,
        'available_referees': available_referees,
        'today_matches': today_matches,
        'deadline_today': deadline_today,
        'deadline_tomorrow': deadline_tomorrow,
        'deadline_week': deadline_week,
    })
@login_required
def submit_comprehensive_report(request, match_id):
    """
    View for referees to submit comprehensive match reports
    """
    # Get the match
    match = get_object_or_404(Match, id=match_id)
    
    # Check if the current user has a referee profile
    try:
        referee = request.user.referee_profile
    except AttributeError:
        messages.error(request, "You don't have a referee profile.")
        return redirect('referees:referee_dashboard')
    
    # Only center referee may submit the comprehensive report.
    # Accept either Match.referee or officials.main_referee if officials object exists.
    is_center_ref = False
    if match.referee and match.referee == referee:
        is_center_ref = True
    else:
        officials = getattr(match, 'officials', None)
        if officials and getattr(officials, 'main_referee', None) == referee:
            is_center_ref = True
    # If not center referee, allow view-only on GET, block POST submissions
    can_edit = bool(is_center_ref)
    
    # Get or create match report - match_id is unique so only use match
    try:
        match_report = MatchReport.objects.get(match=match)
        created = False
    except MatchReport.DoesNotExist:
        match_report = MatchReport.objects.create(
            match=match,
            referee=referee,
            status='draft',
            submitted_at=None
        )
        created = True

    result_form = MatchReportForm(instance=match_report)
    score_form = MatchScoreForm(instance=match)
    GoalFormSetCls = inlineformset_factory(Match, MatchGoal, form=MatchGoalForm, extra=1, can_delete=True, fk_name='match')
    CardFormSetCls = inlineformset_factory(Match, Caution, form=CautionForm, extra=1, can_delete=True, fk_name='match')
    StartingLineupFormSetCls = inlineformset_factory(
        Match, StartingLineup,
        extra=0,
        can_delete=True,
        fk_name='match',
        fields=['player', 'jersey_number', 'position', 'team']
    )
    ReserveFormSetCls = inlineformset_factory(
        Match, ReservePlayer,
        extra=0,
        can_delete=True,
        fk_name='match',
        fields=['player', 'jersey_number', 'team']
    )
    
    # Auto-populate from approved matchday squads if not already populated
    existing_starting = StartingLineup.objects.filter(match=match).count()
    
    if existing_starting == 0:  # Only auto-populate if empty
        # Get approved squads
        home_squad = MatchdaySquad.objects.filter(
            match=match, 
            team=match.home_team, 
            status='approved'
        ).first()
        away_squad = MatchdaySquad.objects.filter(
            match=match, 
            team=match.away_team, 
            status='approved'
        ).first()
        
        # Populate starting lineup and reserves from approved squads
        if home_squad:
            for squad_player in home_squad.squad_players.filter(is_starting=True):
                StartingLineup.objects.get_or_create(
                    match=match,
                    team=home_squad.team,
                    player=squad_player.player,
                    defaults={
                        'jersey_number': squad_player.player.jersey_number,
                        'position': squad_player.player.get_position_display()
                    }
                )
            for squad_player in home_squad.squad_players.filter(is_starting=False):
                ReservePlayer.objects.get_or_create(
                    match=match,
                    team=home_squad.team,
                    player=squad_player.player,
                    defaults={
                        'jersey_number': squad_player.player.jersey_number
                    }
                )
        
        if away_squad:
            for squad_player in away_squad.squad_players.filter(is_starting=True):
                StartingLineup.objects.get_or_create(
                    match=match,
                    team=away_squad.team,
                    player=squad_player.player,
                    defaults={
                        'jersey_number': squad_player.player.jersey_number,
                        'position': squad_player.player.get_position_display()
                    }
                )
            for squad_player in away_squad.squad_players.filter(is_starting=False):
                ReservePlayer.objects.get_or_create(
                    match=match,
                    team=away_squad.team,
                    player=squad_player.player,
                    defaults={
                        'jersey_number': squad_player.player.jersey_number
                    }
                )
    
    goal_formset = GoalFormSetCls(instance=match, form_kwargs={'match': match}, prefix='goal')
    card_formset = CardFormSetCls(instance=match, form_kwargs={'match': match}, prefix='card')
    starting_formset = StartingLineupFormSetCls(instance=match, prefix='starting')
    reserve_formset = ReserveFormSetCls(instance=match, prefix='reserve')
    goalkeeper_form = None  # Add if you have a GoalkeeperForm
    topscorer_form = None   # Add if you have a TopScorerForm

    if request.method == 'POST':
        print("POST request received!")
        print("POST data keys:", list(request.POST.keys()))
        print("Submit button:", request.POST.get('save_draft'), request.POST.get('submit_report'))
        
        if not can_edit:
            messages.error(request, "You are not allowed to submit this report.")
            return redirect('referees:referee_dashboard')
        result_form = MatchReportForm(request.POST, instance=match_report)
        score_form = MatchScoreForm(request.POST, instance=match)
        goal_formset = GoalFormSetCls(request.POST, instance=match, form_kwargs={'match': match}, prefix='goal')
        card_formset = CardFormSetCls(request.POST, instance=match, form_kwargs={'match': match}, prefix='card')
        starting_formset = StartingLineupFormSetCls(request.POST, instance=match, prefix='starting')
        reserve_formset = ReserveFormSetCls(request.POST, instance=match, prefix='reserve')

        # Filter player choices by team for each bound form
        from teams.models import Player
        def _filter_players(fs):
            for i, f in enumerate(fs.forms):
                team_val = f.data.get(f"{fs.prefix}-{i}-team") or getattr(f.instance, 'team_id', None)
                if team_val:
                    f.fields['player'].queryset = Player.objects.filter(team_id=team_val)
        _filter_players(starting_formset)
        _filter_players(reserve_formset)

        # Handle full-time score submission explicitly
        if 'submit_score' in request.POST and score_form.is_valid():
            score_form.save()
            # Mark match completed
            Match.objects.filter(pk=match.pk).update(status='completed')
            # Update simple league table stats for both teams
            from matches.models import LeagueTable
            home = match.home_team
            away = match.away_team
            hs = match.home_score
            ascore = match.away_score
            home_table, _ = LeagueTable.objects.get_or_create(team=home, zone=home.zone)
            away_table, _ = LeagueTable.objects.get_or_create(team=away, zone=away.zone)
            # Only apply if not previously completed (basic safeguard)
            home_table.matches_played += 1
            away_table.matches_played += 1
            home_table.goals_for += hs
            home_table.goals_against += ascore
            away_table.goals_for += ascore
            away_table.goals_against += hs
            if hs > ascore:
                home_table.wins += 1
                away_table.losses += 1
            elif hs < ascore:
                away_table.wins += 1
                home_table.losses += 1
            else:
                home_table.draws += 1
                away_table.draws += 1
            home_table.goal_difference = home_table.goals_for - home_table.goals_against
            away_table.goal_difference = away_table.goals_for - away_table.goals_against
            home_table.calculate_points()
            away_table.calculate_points()
            home_table.save()
            away_table.save()
            messages.success(request, "Full-time score saved and league table updated.")
            return redirect('referees:submit_comprehensive_report', match_id=match.id)

        # Save comprehensive details (draft or final submission)
        # Debug validation errors
        if not result_form.is_valid():
            print("Result form errors:", result_form.errors)
        if not goal_formset.is_valid():
            print("Goal formset errors:", goal_formset.errors)
            print("Goal formset non_form_errors:", goal_formset.non_form_errors())
        if not card_formset.is_valid():
            print("Card formset errors:", card_formset.errors)
            print("Card formset non_form_errors:", card_formset.non_form_errors())
        if not starting_formset.is_valid():
            print("Starting formset errors:", starting_formset.errors)
            print("Starting formset non_form_errors:", starting_formset.non_form_errors())
        if not reserve_formset.is_valid():
            print("Reserve formset errors:", reserve_formset.errors)
            print("Reserve formset non_form_errors:", reserve_formset.non_form_errors())

        if result_form.is_valid() and goal_formset.is_valid() and card_formset.is_valid() and starting_formset.is_valid() and reserve_formset.is_valid():
            result_form.save()
            instances = goal_formset.save()
            card_formset.save()
            starting_formset.save()
            reserve_formset.save()

            # Update top scorers (PlayerStatistic goals) based on saved goals
            from fkf_league.statistics.models import PlayerStatistic
            season = str(timezone.now().year)
            for g in instances:
                ps, _ = PlayerStatistic.objects.get_or_create(player=g.player, season=season)
                ps.goals += 1
                ps.save()
                # Also update simple player tally
                g.player.goals_scored = (g.player.goals_scored or 0) + 1
                g.player.save(update_fields=['goals_scored'])

            # If saving as draft, keep status and redirect back to form
            if 'save_draft' in request.POST:
                match_report.status = 'draft'
                match_report.submitted_at = None
                match_report.save(update_fields=['status', 'submitted_at'])
                messages.success(request, "Draft saved. You can continue adding entries.")
                return redirect('referees:submit_comprehensive_report', match_id=match.id)

            # Final submission
            if 'submit_report' in request.POST:
                match_report.status = 'submitted'
                match_report.submitted_at = timezone.now()
                match_report.save(update_fields=['status', 'submitted_at'])
                messages.success(request, f"Comprehensive report submitted for {match.home_team} vs {match.away_team}. Awaiting manager approval.")
                return redirect('referees:referee_dashboard')
        else:
            # Form validation failed - show errors
            messages.error(request, "Please correct the errors below and try again.")
            # Add specific error messages
            if not result_form.is_valid():
                for field, errors in result_form.errors.items():
                    for error in errors:
                        messages.error(request, f"Report {field}: {error}")
            if not goal_formset.is_valid():
                messages.error(request, "Please check the goals section for errors.")
            if not card_formset.is_valid():
                messages.error(request, "Please check the cautions section for errors.")
            if not starting_formset.is_valid():
                messages.error(request, "Please check the starting lineup section for errors.")
            if not reserve_formset.is_valid():
                messages.error(request, "Please check the reserves section for errors.")

    # On GET, restrict existing forms' player choices to their team (if any)
    from teams.models import Player
    for i, f in enumerate(starting_formset.forms):
        team_val = getattr(f.instance, 'team_id', None)
        if team_val:
            f.fields['player'].queryset = Player.objects.filter(team_id=team_val)
    for i, f in enumerate(reserve_formset.forms):
        team_val = getattr(f.instance, 'team_id', None)
        if team_val:
            f.fields['player'].queryset = Player.objects.filter(team_id=team_val)

    # Team-specific empty forms for client adds (properly prefixed)
    StartingFormClass = StartingLineupFormSetCls.form
    ReserveFormClass = ReserveFormSetCls.form
    starting_empty_home = StartingFormClass(prefix='starting-__prefix__')
    starting_empty_home.fields['player'].queryset = Player.objects.filter(team=match.home_team)
    starting_empty_away = StartingFormClass(prefix='starting-__prefix__')
    starting_empty_away.fields['player'].queryset = Player.objects.filter(team=match.away_team)
    reserve_empty_home = ReserveFormClass(prefix='reserve-__prefix__')
    reserve_empty_home.fields['player'].queryset = Player.objects.filter(team=match.home_team)
    reserve_empty_away = ReserveFormClass(prefix='reserve-__prefix__')
    reserve_empty_away.fields['player'].queryset = Player.objects.filter(team=match.away_team)

    context = {
        'match': match,
        'referee': referee,
        'confirmed': True,
        'result_form': result_form,
        'score_form': score_form,
        'report_form': result_form,
        'goal_formset': goal_formset,
        'card_formset': card_formset,
        'starting_formset': starting_formset,
        'reserve_formset': reserve_formset,
        'starting_empty_home': starting_empty_home,
        'starting_empty_away': starting_empty_away,
        'reserve_empty_home': reserve_empty_home,
        'reserve_empty_away': reserve_empty_away,
        'goalkeeper_form': goalkeeper_form,
        'topscorer_form': topscorer_form,
        'title': 'Submit Comprehensive Report',
        'can_edit': can_edit,
        'home_players': match.home_team.players.all(),
        'away_players': match.away_team.players.all(),
        'halftime_substitutions': match.substitutions.filter(minute__gte=45, minute__lte=60),
    }
    return render(request, 'referees/match_report_comprehensive.html', context)


# ========== PRE-MATCH MEETING FORM ==========
@login_required
def submit_prematch_form(request, match_id):
    """Submit pre-match meeting form - main referee only, 36 hours before match"""
    from .models import PreMatchMeetingForm
    from .forms import PreMatchMeetingFormForm
    from teams.models import Team
    
    match = get_object_or_404(Match, id=match_id)
    
    # Check if user is the main referee for this match
    try:
        referee = request.user.referee_profile
        officials = match.officials
        
        if officials.main_referee != referee:
            messages.error(request, "Only the main referee can submit the pre-match meeting form.")
            return redirect('referees:referee_dashboard')
            
    except AttributeError:
        messages.error(request, "You don't have a referee profile.")
        return redirect('referees:referee_dashboard')
    except MatchOfficials.DoesNotExist:
        messages.error(request, "No officials have been appointed for this match yet.")
        return redirect('referees:referee_dashboard')
    
    # Check if match is within 36 hours
    time_until_match = match.match_date - timezone.now()
    if time_until_match > timedelta(hours=36):
        hours_remaining = int(time_until_match.total_seconds() / 3600)
        messages.warning(request, f"Pre-match form can only be filled within 36 hours of the match. {hours_remaining} hours remaining.")
        return redirect('referees:referee_dashboard')
    
    # Check if match has already occurred
    if timezone.now() >= match.match_date:
        messages.error(request, "Cannot fill pre-match form after the match has started.")
        return redirect('referees:referee_dashboard')
    
    # Get or create pre-match form
    prematch_form, created = PreMatchMeetingForm.objects.get_or_create(
        match=match,
        defaults={
            'referee': referee,
            'match_date': match.match_date.date() if hasattr(match.match_date, 'date') else match.match_date,
            'match_number': f"{match.id}",
            'home_team': match.home_team.team_name,
            'away_team': match.away_team.team_name,
            'venue': match.venue or '',
        }
    )
    
    # Check if form is locked
    if prematch_form.is_locked:
        messages.error(request, "This form is locked and cannot be edited.")
        return redirect('referees:view_prematch_form', form_id=prematch_form.id)
    
    if request.method == 'POST':
        form = PreMatchMeetingFormForm(request.POST, instance=prematch_form)
        
        if form.is_valid():
            prematch = form.save(commit=False)
            
            # Auto-fill match details
            prematch.match_date = match.match_date.date() if hasattr(match.match_date, 'date') else match.match_date
            prematch.home_team = match.home_team.team_name
            prematch.away_team = match.away_team.team_name
            prematch.venue = match.venue or ''
            
            # Auto-fill officials from MatchOfficials
            if hasattr(match, 'officials'):
                offs = match.officials
                if offs.match_commissioner:
                    prematch.match_commissioner_name = offs.match_commissioner.full_name
                    prematch.match_commissioner_license = offs.match_commissioner.fkf_number
                    prematch.match_commissioner_mobile = offs.match_commissioner.phone_number
                if offs.main_referee:
                    prematch.centre_referee_name = offs.main_referee.full_name
                    prematch.centre_referee_license = offs.main_referee.fkf_number
                    prematch.centre_referee_mobile = offs.main_referee.phone_number
                if offs.assistant_1:
                    prematch.asst1_referee_name = offs.assistant_1.full_name
                    prematch.asst1_referee_license = offs.assistant_1.fkf_number
                    prematch.asst1_referee_mobile = offs.assistant_1.phone_number
                if offs.assistant_2:
                    prematch.asst2_referee_name = offs.assistant_2.full_name
                    prematch.asst2_referee_license = offs.assistant_2.fkf_number
                    prematch.asst2_referee_mobile = offs.assistant_2.phone_number
                if offs.fourth_official:
                    prematch.fourth_official_name = offs.fourth_official.full_name
                    prematch.fourth_official_license = offs.fourth_official.fkf_number
                    prematch.fourth_official_mobile = offs.fourth_official.phone_number
            
            # Auto-fill kit colors from team
            home_team_obj = Team.objects.get(id=match.home_team.id)
            away_team_obj = Team.objects.get(id=match.away_team.id)
            
            # Home team kits (using home colors)
            prematch.home_gk_shirt_color = home_team_obj.gk_home_jersey_color
            prematch.home_official_shirt = home_team_obj.home_jersey_color
            
            # Away team kits (using away colors)
            prematch.away_gk_shirt_color = away_team_obj.gk_away_jersey_color
            prematch.away_official_shirt = away_team_obj.away_jersey_color
            
            # Submit or save as draft
            if 'submit' in request.POST:
                prematch.status = 'submitted'
                prematch.submitted_at = timezone.now()
                messages.success(request, "✅ Pre-Match Meeting Form completed and submitted successfully! Awaiting admin approval.")
            else:
                prematch.status = 'pending'
                messages.success(request, "💾 Pre-Match Form saved as draft.")
            
            prematch.save()
            return redirect('referees:referee_dashboard')
    else:
        # Auto-fill form on GET
        if created or not prematch_form.submitted_at:
            # Auto-fill officials
            if hasattr(match, 'officials'):
                offs = match.officials
                if offs.match_commissioner:
                    prematch_form.match_commissioner_name = offs.match_commissioner.full_name
                    prematch_form.match_commissioner_license = offs.match_commissioner.fkf_number
                    prematch_form.match_commissioner_mobile = offs.match_commissioner.phone_number
                if offs.main_referee:
                    prematch_form.centre_referee_name = offs.main_referee.full_name
                    prematch_form.centre_referee_license = offs.main_referee.fkf_number
                    prematch_form.centre_referee_mobile = offs.main_referee.phone_number
                if offs.assistant_1:
                    prematch_form.asst1_referee_name = offs.assistant_1.full_name
                    prematch_form.asst1_referee_license = offs.assistant_1.fkf_number
                    prematch_form.asst1_referee_mobile = offs.assistant_1.phone_number
                if offs.assistant_2:
                    prematch_form.asst2_referee_name = offs.assistant_2.full_name
                    prematch_form.asst2_referee_license = offs.assistant_2.fkf_number
                    prematch_form.asst2_referee_mobile = offs.assistant_2.phone_number
                if offs.fourth_official:
                    prematch_form.fourth_official_name = offs.fourth_official.full_name
                    prematch_form.fourth_official_license = offs.fourth_official.fkf_number
                    prematch_form.fourth_official_mobile = offs.fourth_official.phone_number
            
            # Auto-fill kit colors
            home_team_obj = Team.objects.get(id=match.home_team.id)
            away_team_obj = Team.objects.get(id=match.away_team.id)
            
            prematch_form.home_gk_shirt_color = home_team_obj.gk_home_jersey_color
            prematch_form.home_official_shirt = home_team_obj.home_jersey_color
            
            prematch_form.away_gk_shirt_color = away_team_obj.gk_away_jersey_color
            prematch_form.away_official_shirt = away_team_obj.away_jersey_color
        
        form = PreMatchMeetingFormForm(instance=prematch_form)
    
    context = {
        'form': form,
        'match': match,
        'prematch_form': prematch_form,
        'is_locked': prematch_form.is_locked,
    }
    return render(request, 'referees/submit_prematch_form.html', context)


@login_required
def view_prematch_form(request, form_id):
    """View submitted pre-match meeting form"""
    from .models import PreMatchMeetingForm
    
    prematch_form = get_object_or_404(PreMatchMeetingForm, id=form_id)
    
    # Check access permissions
    is_main_referee = False
    is_manager = request.user.is_staff or request.user.groups.filter(name='Referees Manager').exists()
    is_admin = request.user.is_staff
    
    try:
        referee = request.user.referee_profile
        if prematch_form.match.officials.main_referee == referee:
            is_main_referee = True
    except:
        pass
    
    if not (is_main_referee or is_manager or is_admin):
        messages.error(request, "You don't have permission to view this form.")
        return redirect('referees:referee_dashboard')
    
    context = {
        'prematch_form': prematch_form,
        'match': prematch_form.match,
        'is_main_referee': is_main_referee,
        'is_manager': is_manager,
        'is_admin': is_admin,
    }
    return render(request, 'referees/view_prematch_form.html', context)


@login_required
@permission_required('referees.approve_prematch_form_admin', raise_exception=True)
def admin_approve_prematch_form(request, form_id):
    """Admin approves pre-match meeting form"""
    from .models import PreMatchMeetingForm
    
    prematch_form = get_object_or_404(PreMatchMeetingForm, id=form_id)
    
    if prematch_form.status != 'submitted':
        messages.error(request, "This form is not awaiting admin approval.")
        return redirect('referees:view_prematch_form', form_id=form_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            prematch_form.status = 'admin_approved'
            prematch_form.admin_approved_at = timezone.now()
            prematch_form.admin_approved_by = request.user
            prematch_form.save()
            messages.success(request, "Pre-match form approved by admin. Awaiting referee manager approval.")
        
        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '')
            prematch_form.status = 'rejected'
            prematch_form.admin_rejection_reason = reason
            prematch_form.save()
            messages.success(request, "Pre-match form rejected.")
        
        return redirect('referees:view_prematch_form', form_id=form_id)
    
    context = {
        'prematch_form': prematch_form,
        'match': prematch_form.match,
    }
    return render(request, 'referees/admin_approve_prematch_form.html', context)


@login_required
@permission_required('referees.approve_prematch_form_manager', raise_exception=True)
def manager_approve_prematch_form(request, form_id):
    """Referee Manager approves pre-match meeting form"""
    from .models import PreMatchMeetingForm
    
    prematch_form = get_object_or_404(PreMatchMeetingForm, id=form_id)
    
    if prematch_form.status != 'admin_approved':
        messages.error(request, "This form is not awaiting manager approval.")
        return redirect('referees:view_prematch_form', form_id=form_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            prematch_form.status = 'approved'
            prematch_form.manager_approved_at = timezone.now()
            prematch_form.manager_approved_by = request.user
            prematch_form.save()
            messages.success(request, "Pre-match form fully approved.")
        
        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '')
            prematch_form.status = 'rejected'
            prematch_form.manager_rejection_reason = reason
            prematch_form.save()
            messages.success(request, "Pre-match form rejected.")
        
        return redirect('referees:view_prematch_form', form_id=form_id)
    
    context = {
        'prematch_form': prematch_form,
        'match': prematch_form.match,
    }
    return render(request, 'referees/manager_approve_prematch_form.html', context)


@login_required
def pending_prematch_forms_admin(request):
    """List of pre-match forms pending admin approval"""
    from .models import PreMatchMeetingForm
    
    if not request.user.is_staff:
        messages.error(request, "Access denied.")
        return redirect('referees:referee_dashboard')
    
    forms = PreMatchMeetingForm.objects.filter(status='submitted').order_by('-match__match_date')
    
    context = {
        'forms': forms,
    }
    return render(request, 'referees/pending_prematch_forms_admin.html', context)


@login_required
def pending_prematch_forms_manager(request):
    """List of pre-match forms pending manager approval"""
    from .models import PreMatchMeetingForm
    
    if not (request.user.is_staff or request.user.groups.filter(name='Referees Manager').exists()):
        messages.error(request, "Access denied.")
        return redirect('referees:referee_dashboard')
    
    forms = PreMatchMeetingForm.objects.filter(status='admin_approved').order_by('-match__match_date')
    
    context = {
        'forms': forms,
    }
    return render(request, 'referees/pending_prematch_forms_manager.html', context)


# ========== MANAGER: APPROVE/REJECT REPORTS ==========
@login_required
def pending_reports(request):
    """List of reports pending approval (Manager only)"""
    # Check if user is a referees manager or admin
    if not (request.user.is_staff or request.user.groups.filter(name='Referees Manager').exists()):
        messages.error(request, "Only managers can access this page.")
        return redirect('referees:referee_dashboard')
    
    # Get all submitted reports
    pending = MatchReport.objects.filter(status='submitted').select_related('match', 'referee', 'approved_by')
    approved_reports = MatchReport.objects.filter(status='approved').select_related('match', 'referee', 'approved_by')
    rejected_reports = MatchReport.objects.filter(status='rejected').select_related('match', 'referee', 'approved_by')
    
    context = {
        'pending_reports': pending,
        'approved_reports': approved_reports,
        'rejected_reports': rejected_reports,
        'pending_count': pending.count(),
        'approved_count': approved_reports.count(),
        'rejected_count': rejected_reports.count(),
        'title': 'Pending Match Reports'
    }
    return render(request, 'referees/manager/pending_reports.html', context)


@login_required
def approve_report(request, report_id):
    """Approve a match report (Manager only)"""
    if not (request.user.is_staff or request.user.groups.filter(name='Referees Manager').exists()):
        messages.error(request, "Only managers can approve reports.")
        return redirect('referees:referee_dashboard')
    
    report = get_object_or_404(MatchReport, id=report_id)
    
    if request.method == 'POST':
        report.status = 'approved'
        report.approved_by = request.user
        report.approved_at = timezone.now()
        report.save()
        
        messages.success(request, f"Report for {report.match} has been approved!")
        return redirect('referees:pending_reports')
    
    context = {
        'report': report,
        'match': report.match,
        'title': f'Approve Report - {report.match}'
    }
    return render(request, 'referees/manager/approve_report.html', context)


@login_required
def reject_report(request, report_id):
    """Reject a match report (Manager only)"""
    if not (request.user.is_staff or request.user.groups.filter(name='Referees Manager').exists()):
        messages.error(request, "Only managers can reject reports.")
        return redirect('referees:referee_dashboard')
    
    report = get_object_or_404(MatchReport, id=report_id)
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        report.status = 'rejected'
        report.approved_by = request.user
        report.approved_at = timezone.now()
        report.save()
        
        # Add rejection reason as a message to the referee
        messages.info(
            request,
            f"Report for {report.match} has been rejected. Reason: {rejection_reason}"
        )
        
        return redirect('referees:pending_reports')
    
    context = {
        'report': report,
        'match': report.match,
        'title': f'Reject Report - {report.match}'
    }
    return render(request, 'referees/manager/reject_report.html', context)


@login_required
def report_detail_view(request, report_id):
    """View report details (read-only for managers)"""
    report = get_object_or_404(MatchReport, id=report_id)
    match = report.match
    
    # Check permissions
    is_manager = request.user.is_staff or request.user.groups.filter(name='Referees Manager').exists()
    is_referee = hasattr(request.user, 'referee_profile') and request.user.referee_profile == report.referee
    
    if not (is_manager or is_referee):
        messages.error(request, "You don't have permission to view this report.")
        return redirect('referees:referee_dashboard')
    
    goals = match.matchgoal_set.all()
    cautions = match.caution_set.all()
    
    context = {
        'report': report,
        'match': match,
        'goals': goals,
        'cautions': cautions,
        'is_manager': is_manager,
        'is_referee': is_referee,
        'title': f'Report - {match}'
    }
    return render(request, 'referees/manager/report_detail.html', context)


@login_required
def submit_match_report(request, match_id):
    """
    View for referees to submit quick match reports
    """
    # Get the match
    match = get_object_or_404(Match, id=match_id)
    
    # Check if the current user has a referee profile
    try:
        referee = request.user.referee_profile
    except AttributeError:
        messages.error(request, "You don't have a referee profile.")
        return redirect('referees:referee_dashboard')
    
    # Check if match has officials
    if not hasattr(match, 'officials'):
        messages.error(request, "No officials appointed for this match.")
        return redirect('referees:referee_dashboard')
    
    officials = match.officials
    
    # Check if the current referee is appointed to this match
    is_appointed = False
    confirmed = False
    
    # Check all official positions
    if officials.main_referee == referee:
        is_appointed = True
        confirmed = officials.main_confirmed
    elif officials.assistant_1 == referee:
        is_appointed = True
        confirmed = officials.ar1_confirmed
    elif officials.assistant_2 == referee:
        is_appointed = True
        confirmed = officials.ar2_confirmed
    elif officials.fourth_official == referee:
        is_appointed = True
        confirmed = officials.fourth_confirmed
    
    if not is_appointed:
        messages.error(request, "You are not appointed to this match.")
        return redirect('referees:referee_dashboard')
    
    if not confirmed:
        messages.error(request, "You need to confirm your appointment before submitting reports.")
        return redirect('referees:confirm_appointment', match_id=match.id)
    
    # Get or create match report
    match_report, created = MatchReport.objects.get_or_create(
        match=match,
        referee=referee,
        defaults={
            'status': 'draft',
            'submitted_at': None
        }
    )

    # Prepare forms and formsets for flash report
    result_form = MatchReportForm(instance=match_report)
    score_form = MatchScoreForm(instance=match)
    
    GoalFormSet = inlineformset_factory(Match, MatchGoal, form=MatchGoalForm, extra=1, fk_name='match')
    goal_formset = GoalFormSet(instance=match, prefix='matchgoal_set', form_kwargs={'match': match})
    
    CardFormSet = inlineformset_factory(Match, Caution, form=CautionForm, extra=1, fk_name='match')
    card_formset = CardFormSet(instance=match, prefix='caution_set', form_kwargs={'match': match})
    
    goalkeeper_form = None  # Add if you have a GoalkeeperForm
    topscorer_form = None   # Add if you have a TopScorerForm

    if request.method == 'POST':
        result_form = MatchReportForm(request.POST, instance=match_report)
        score_form = MatchScoreForm(request.POST, instance=match)
        goal_formset = GoalFormSet(request.POST, instance=match, prefix='matchgoal_set', form_kwargs={'match': match})
        card_formset = CardFormSet(request.POST, instance=match, prefix='caution_set', form_kwargs={'match': match})
        # Add goalkeeper_form and topscorer_form POST handling if present
        if result_form.is_valid() and score_form.is_valid() and goal_formset.is_valid() and card_formset.is_valid():
            result_form.save()
            score_form.save()
            goal_formset.save()
            card_formset.save()
            messages.success(request, f"✅ Flash Report submitted successfully for {match.home_team} vs {match.away_team}!")
            return redirect('referees:referee_dashboard')
        else:
            messages.error(request, "❌ Please correct the errors in the form before submitting.")

    context = {
        'match': match,
        'referee': referee,
        'officials': officials,
        'result_form': result_form,
        'score_form': score_form,
        'goal_formset': goal_formset,
        'card_formset': card_formset,
        'goalkeeper_form': goalkeeper_form,
        'topscorer_form': topscorer_form,
        'title': 'Submit Flash Match Report',
        'home_players': match.home_team.players.all(),
        'away_players': match.away_team.players.all(),
    }
    return render(request, 'referees/flash_report.html', context)




@login_required
@require_http_methods(["GET", "POST"])
def generate_weekly_report(request):
    """
    Generate weekly report with different views based on user role
    """
    try:
        # Get date parameters
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
            except:
                data = request.POST
        else:
            data = request.GET
        
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        
        if start_date_str and end_date_str:
            # Use provided dates
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        elif data.get('period') == 'dec_22_31_2025':
            # Specific: Dec 22-31, 2025
            from datetime import date
            start_date = date(2025, 12, 22)
            end_date = date(2025, 12, 31)
        else:
            # Default: Last 30 days
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        
        from matches.models import Match
        
        # Check user role
        user_groups = [g.name for g in request.user.groups.all()]
        
        # Check if user is in Referee Managers group
        is_manager = any('manager' in g.name.lower() for g in request.user.groups.all()) or \
                    request.user.groups.filter(name='Referee Managers').exists()
        
        is_admin = request.user.is_staff or request.user.is_superuser
        
        # Check if user has a referee profile
        has_referee_profile = hasattr(request.user, 'referee_profile')
        
        # Convert dates to timezone-aware datetime objects for filtering
        from datetime import datetime as dt, time
        start_datetime = timezone.make_aware(dt.combine(start_date, time.min))
        end_datetime = timezone.make_aware(dt.combine(end_date, time.max))
        
        # Get matches in date range
        matches = Match.objects.filter(
            match_date__gte=start_datetime,
            match_date__lte=end_datetime
        ).select_related(
            'home_team', 'away_team', 'zone',
            'referee', 'assistant_referee_1', 'assistant_referee_2'
        ).order_by('match_date', 'kickoff_time')
        
        if is_manager or is_admin:
            # MANAGER VIEW: Show comprehensive overview
            
            # Helper function to get official's name
            def get_official_name(official):
                if not official:
                    return None
                if hasattr(official, 'user'):
                    return official.user.get_full_name()
                elif hasattr(official, 'get_full_name'):
                    return official.get_full_name()
                return str(official)
            
            # Analyze match status
            total_matches = matches.count()
            fully_assigned = 0
            partially_assigned = 0
            unassigned = 0
            
            matches_data = []
            for match in matches:
                # Check assignment status
                has_referee = match.referee is not None
                has_ar1 = match.assistant_referee_1 is not None
                has_ar2 = match.assistant_referee_2 is not None
                
                if has_referee and has_ar1 and has_ar2:
                    fully_assigned += 1
                    assignment_status = 'fully_assigned'
                elif has_referee or has_ar1 or has_ar2:
                    partially_assigned += 1
                    assignment_status = 'partially_assigned'
                else:
                    unassigned += 1
                    assignment_status = 'unassigned'
                
                # Get team names safely
                home_team_name = match.home_team.team_name if match.home_team else 'Unknown'
                away_team_name = match.away_team.team_name if match.away_team else 'Unknown'
                
                # Format date safely
                if hasattr(match.match_date, 'isoformat'):
                    match_date_str = match.match_date.isoformat()
                else:
                    match_date_str = str(match.match_date)
                
                # Format kickoff_time (it's a CharField, not TimeField)
                kickoff_time_str = ''
                if match.kickoff_time:
                    # Clean up the time string
                    time_str = str(match.kickoff_time).strip()
                    if time_str:
                        # Remove any seconds if present (HH:MM:SS -> HH:MM)
                        if ':' in time_str:
                            parts = time_str.split(':')
                            if len(parts) >= 2:
                                kickoff_time_str = f"{parts[0]}:{parts[1]}"
                
                match_info = {
                    'id': match.id,
                    'date': match_date_str,
                    'time': kickoff_time_str,
                    'venue': match.venue if hasattr(match, 'venue') else '',
                    'venue_details': match.venue_details if hasattr(match, 'venue_details') else '',
                    'zone': match.zone.name if match.zone else '',
                    'teams': f"{home_team_name} vs {away_team_name}",
                    'home_team': {
                        'id': match.home_team.id if match.home_team else None,
                        'name': home_team_name,
                        'code': match.home_team.team_code if match.home_team else ''
                    },
                    'away_team': {
                        'id': match.away_team.id if match.away_team else None,
                        'name': away_team_name,
                        'code': match.away_team.team_code if match.away_team else ''
                    },
                    'assignment_status': assignment_status,
                    'officials': {
                        'referee': get_official_name(match.referee),
                        'assistant_referee_1': get_official_name(match.assistant_referee_1),
                        'assistant_referee_2': get_official_name(match.assistant_referee_2)
                    },
                    'scores': {
                        'home': match.home_score,
                        'away': match.away_score
                    },
                    'status': match.status,
                    'needs_attention': assignment_status in ['partially_assigned', 'unassigned']
                }
                matches_data.append(match_info)
            
            # Count officials needed
            needs_referee = sum(1 for m in matches if m.referee is None)
            needs_ar1 = sum(1 for m in matches if m.assistant_referee_1 is None)
            needs_ar2 = sum(1 for m in matches if m.assistant_referee_2 is None)
            
            report_data = {
                'report_type': 'manager_overview',
                'user_role': 'manager' if is_manager else 'admin',
                'user_info': {
                    'username': request.user.username,
                    'groups': user_groups,
                    'is_staff': request.user.is_staff,
                    'is_superuser': request.user.is_superuser
                },
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat(),
                    'days': (end_date - start_date).days + 1
                },
                'summary': {
                    'total_matches': total_matches,
                    'fully_assigned': fully_assigned,
                    'partially_assigned': partially_assigned,
                    'unassigned': unassigned,
                    'officials_needed': {
                        'referees': needs_referee,
                        'assistant_referee_1': needs_ar1,
                        'assistant_referee_2': needs_ar2,
                        'total': needs_referee + needs_ar1 + needs_ar2
                    }
                },
                'matches': matches_data,
                'matches_needing_attention': [
                    m for m in matches_data if m['needs_attention']
                ]
            }
            
        elif has_referee_profile:
            # REFEREE VIEW: Show only their assignments
            referee_profile = request.user.referee_profile
            
            # Filter matches for this referee
            referee_matches = matches.filter(
                models.Q(referee=referee_profile) |
                models.Q(assistant_referee_1=referee_profile) |
                models.Q(assistant_referee_2=referee_profile)
            )
            
            # Helper function to get official's name
            def get_official_name(official):
                if not official:
                    return None
                if hasattr(official, 'user'):
                    return official.user.get_full_name()
                elif hasattr(official, 'get_full_name'):
                    return official.get_full_name()
                return str(official)
            
            assignments = []
            for match in referee_matches:
                # Determine the referee's role in this match
                if match.referee == referee_profile:
                    role = 'referee'
                elif match.assistant_referee_1 == referee_profile:
                    role = 'assistant_referee_1'
                else:
                    role = 'assistant_referee_2'
                
                # Get other officials (excluding self)
                other_officials = {}
                if match.referee and match.referee != referee_profile:
                    other_officials['referee'] = get_official_name(match.referee)
                if match.assistant_referee_1 and match.assistant_referee_1 != referee_profile:
                    other_officials['assistant_referee_1'] = get_official_name(match.assistant_referee_1)
                if match.assistant_referee_2 and match.assistant_referee_2 != referee_profile:
                    other_officials['assistant_referee_2'] = get_official_name(match.assistant_referee_2)
                
                # Get team names safely
                home_team_name = match.home_team.team_name if match.home_team else 'Unknown'
                away_team_name = match.away_team.team_name if match.away_team else 'Unknown'
                
                # Format kickoff_time
                kickoff_time_str = ''
                if match.kickoff_time:
                    time_str = str(match.kickoff_time).strip()
                    if time_str and ':' in time_str:
                        parts = time_str.split(':')
                        if len(parts) >= 2:
                            kickoff_time_str = f"{parts[0]}:{parts[1]}"
                
                assignment_info = {
                    'id': match.id,
                    'date': match.match_date.isoformat() if hasattr(match.match_date, 'isoformat') else str(match.match_date),
                    'time': kickoff_time_str,
                    'venue': match.venue if hasattr(match, 'venue') else '',
                    'venue_details': match.venue_details if hasattr(match, 'venue_details') else '',
                    'zone': match.zone.name if match.zone else '',
                    'teams': f"{home_team_name} vs {away_team_name}",
                    'your_role': role.replace('_', ' ').title(),
                    'other_officials': other_officials,
                    'scores': {
                        'home': match.home_score,
                        'away': match.away_score
                    },
                    'status': match.status,
                    'match_status': 'completed' if match.status == 'completed' else 'upcoming'
                }
                assignments.append(assignment_info)
            
            # Count by role
            as_referee = sum(1 for a in assignments if a['your_role'] == 'Referee')
            as_ar1 = sum(1 for a in assignments if a['your_role'] == 'Assistant Referee 1')
            as_ar2 = sum(1 for a in assignments if a['your_role'] == 'Assistant Referee 2')
            
            report_data = {
                'report_type': 'referee_assignments',
                'user_role': 'referee',
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'referee_name': request.user.get_full_name(),
                'referee_id': referee_profile.id,
                'summary': {
                    'total_assignments': len(assignments),
                    'as_referee': as_referee,
                    'as_assistant_referee_1': as_ar1,
                    'as_assistant_referee_2': as_ar2,
                    'completed': sum(1 for a in assignments if a['match_status'] == 'completed'),
                    'upcoming': sum(1 for a in assignments if a['match_status'] == 'upcoming')
                },
                'assignments': assignments
            }
            
        else:
            # REGULAR USER VIEW
            matches_data = []
            for match in matches:
                home_team_name = match.home_team.team_name if match.home_team else 'Unknown'
                away_team_name = match.away_team.team_name if match.away_team else 'Unknown'
                
                # Format kickoff_time
                kickoff_time_str = ''
                if match.kickoff_time:
                    time_str = str(match.kickoff_time).strip()
                    if time_str and ':' in time_str:
                        parts = time_str.split(':')
                        if len(parts) >= 2:
                            kickoff_time_str = f"{parts[0]}:{parts[1]}"
                
                match_info = {
                    'id': match.id,
                    'date': match.match_date.isoformat() if hasattr(match.match_date, 'isoformat') else str(match.match_date),
                    'time': kickoff_time_str,
                    'venue': match.venue if hasattr(match, 'venue') else '',
                    'venue_details': match.venue_details if hasattr(match, 'venue_details') else '',
                    'zone': match.zone.name if match.zone else '',
                    'teams': f"{home_team_name} vs {away_team_name}",
                    'scores': {
                        'home': match.home_score,
                        'away': match.away_score
                    },
                    'status': match.status
                }
                matches_data.append(match_info)
            
            report_data = {
                'report_type': 'public_schedule',
                'user_role': 'public',
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'summary': {
                    'total_matches': len(matches_data)
                },
                'matches': matches_data
            }
        
        return JsonResponse(report_data, safe=False)
        
    except Exception as e:
        import traceback
        error_response = {
            'error': str(e),
            'user_type': type(request.user).__name__,
        }
        
        if request.user.is_staff:
            error_response['traceback'] = traceback.format_exc()
            
        return JsonResponse(error_response, status=400)
@login_required
def weekly_report_display(request):
    """HTML display for weekly report - Modern version"""
    return render(request, 'referees/weekly_report_modern.html')


@login_required
def export_appointments_excel(request):
    """Export appointments to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from datetime import datetime as dt, time
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = dt.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = dt.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
    
    # Convert to timezone-aware datetime
    start_datetime = timezone.make_aware(dt.combine(start_date, time.min))
    end_datetime = timezone.make_aware(dt.combine(end_date, time.max))
    
    # Get matches
    matches = Match.objects.filter(
        match_date__gte=start_datetime,
        match_date__lte=end_datetime
    ).select_related(
        'home_team', 'away_team', 'zone',
        'referee', 'assistant_referee_1', 'assistant_referee_2'
    ).order_by('match_date', 'kickoff_time')
    
    # Helper function to get official's name (only approved with user accounts)
    def get_official_name(official):
        if not official:
            return 'VACANT'
        # Only show referee if approved, has user account, and not suspended
        if hasattr(official, 'status') and official.status != 'approved':
            return 'VACANT'
        if hasattr(official, 'user') and official.user:
            return official.user.get_full_name() or f"{official.first_name} {official.last_name}"
        # If no user account, referee is not yet activated
        return 'VACANT'
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Match Officials Report"
    
    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"FKF Meru League - Match Officials Report ({start_date} to {end_date})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Date', 'Time', 'Match', 'Venue', 'Zone', 'Referee', 'AR1', 'AR2']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    row = 4
    for match in matches:
        ws.cell(row=row, column=1).value = match.match_date.strftime('%Y-%m-%d') if match.match_date else ''
        ws.cell(row=row, column=2).value = str(match.kickoff_time) if match.kickoff_time else 'TBD'
        ws.cell(row=row, column=3).value = f"{match.home_team.team_name} vs {match.away_team.team_name}"
        ws.cell(row=row, column=4).value = match.venue or match.venue_details or 'TBD'
        ws.cell(row=row, column=5).value = match.zone.name if match.zone else ''
        ws.cell(row=row, column=6).value = get_official_name(match.referee)
        ws.cell(row=row, column=7).value = get_official_name(match.assistant_referee_1)
        ws.cell(row=row, column=8).value = get_official_name(match.assistant_referee_2)
        
        # Apply borders
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Match_Officials_Report_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_appointments_pdf(request):
    """Export appointments to PDF with MKJ SUPA CUP 11th edition logo"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from django.http import HttpResponse
    from django.conf import settings
    from datetime import datetime as dt, time
    import io
    import os
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = dt.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = dt.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
    
    # Convert to timezone-aware datetime
    start_datetime = timezone.make_aware(dt.combine(start_date, time.min))
    end_datetime = timezone.make_aware(dt.combine(end_date, time.max))
    
    # Get matches
    matches = Match.objects.filter(
        match_date__gte=start_datetime,
        match_date__lte=end_datetime
    ).select_related(
        'home_team', 'away_team', 'zone',
        'referee', 'assistant_referee_1', 'assistant_referee_2'
    ).order_by('match_date', 'kickoff_time')
    
    # Helper function to get official's name (only approved with user accounts)
    def get_official_name(official):
        if not official:
            return 'VACANT'
        # Only show referee if approved, has user account, and not suspended
        if hasattr(official, 'status') and official.status != 'approved':
            return 'VACANT'
        if hasattr(official, 'user') and official.user:
            return official.user.get_full_name() or f"{official.first_name} {official.last_name}"
        # If no user account, referee is not yet activated
        return 'VACANT'
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Create elements list and add MKJ SUPA CUP 11th edition logo
    elements = []
    logo_path = os.path.join(settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else settings.STATIC_ROOT, 'img', 'mkj_supacup_logo_official.jpg')
    if os.path.exists(logo_path):
        try:
            logo = Image(logo_path, width=40*mm, height=40*mm)
            logo.hAlign = 'LEFT'
            elements.append(logo)
            elements.append(Spacer(1, 2*mm))
        except:
            pass  # Logo load failed, continue without it
    
    # Styles
    styles = getSampleStyleSheet()
    edition_style = ParagraphStyle("Edition", parent=styles["Heading2"], fontSize=10, textColor=colors.HexColor("#004D1A"), spaceAfter=6, alignment=1)
    elements.append(Paragraph("⚽ KENYA YOUTH INTERCOUNTY SPORTS ASSOCIATION — 11TH EDITION", edition_style))
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=20,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph(f"FKF Meru League - Match Officials Report<br/>{start_date} to {end_date}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [['Date', 'Time', 'Match', 'Venue', 'Referee', 'AR1', 'AR2']]
    
    for match in matches:
        data.append([
            match.match_date.strftime('%d/%m/%Y') if match.match_date else '',
            str(match.kickoff_time)[:5] if match.kickoff_time else 'TBD',
            f"{match.home_team.team_name}\nvs\n{match.away_team.team_name}",
            match.venue or match.venue_details or 'TBD',
            get_official_name(match.referee),
            get_official_name(match.assistant_referee_1),
            get_official_name(match.assistant_referee_2)
        ])
    
    # Create table
    table = Table(data, colWidths=[0.8*inch, 0.6*inch, 2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Response
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Match_Officials_Report_{start_date}_{end_date}.pdf"'
    return response
@login_required
@permission_required('referees.submit_match_report', raise_exception=True)
def view_report(request, report_id):
    """
    View a specific match report
    """
    # Get the report
    report = get_object_or_404(MatchReport, id=report_id)
    
    # Check if the current user has a referee profile
    try:
        referee = request.user.referee_profile
    except AttributeError:
        messages.error(request, "You don't have a referee profile.")
        return redirect('referees:referee_dashboard')
    
    # Check if the current referee owns this report or is admin
    if report.referee != referee and not request.user.is_staff:
        messages.error(request, "You don't have permission to view this report.")
        return redirect('referees:referee_dashboard')
    
    # Get match details
    match = report.match
    
    context = {
        'report': report,
        'match': match,
        'title': f'Match Report #{report.id}',
        'referee': referee,
        'is_owner': report.referee == referee,
    }
    
    return render(request, 'referees/view_report.html', context)